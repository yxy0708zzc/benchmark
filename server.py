"""
FastAPI 服务主程序

包含：
- 7 个基础数据查询端点（/api/train、/api/station、/api/station/random、
  /api/train/{num}、/api/station/{id}、/api/train/{num}/ticket、/api/routes）
- 出题器端点（update_ticket, auto_generate/previews/clear/confirm, question/*）
- 测试器端点（chat/stream, complete, reset, records）
- 测评器端点（eval/load, eval/verify, eval/results, eval/complete）
- 统计端点（stats/summary, stats/export/json, stats/export/markdown）
- 静态文件服务（前端页面）
"""

import json
import os
import sqlite3
import random
import re
import logging
import asyncio
import threading
from io import BytesIO
from datetime import datetime
from typing import Optional, List, Dict, Any, Tuple
from contextlib import asynccontextmanager

from fastapi import FastAPI, HTTPException, Query, Body, Request, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse, StreamingResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from config import (
    API_CONFIG, QUESTION_CONFIG, ensure_directories,
    RAILWAY_DB_PATH, METADATA_PATH, LOGS_DIR,
    get_question_db_path, TEMPLATES_DIR,
    NL_TEMPERATURE, TEST_TEMPERATURE
)
from database import (
    get_railway_conn,
    get_train_stops, get_routes_between, get_station_trains,
    get_train_tickets, get_train_tickets_filtered,
    create_question_db, reset_question_tables, update_ticket as db_update_ticket,
    delete_train_tickets, remove_train_from_metadata,
    load_metadata, update_question_metadata, get_question_metadata,
    save_metadata,
    validate_train_exists, validate_station_exists,
    resolve_station_name_or_id,
    load_same_train_map,
    TICKET_TABLES
)
from prompts import SYSTEM_PROMPT

# ============================================================
# 工具定义（原 tools/__init__.py，已整合进本文件）
# 定义被测模型可调用的 7 个 OpenAI Tools
#
# 每个工具包含：
# - function 字典（用于传入大模型的 tool 定义）
# - handler 函数（实际执行逻辑，在下方 execute_tool_handler 中分发）
# ============================================================

# ============================================================
# 全局状态：当前题目 ID
# 由测试器在启动时设置
# ============================================================
_current_question_id: Optional[str] = None


def set_current_question(question_id: str):
    """设置当前题目 ID"""
    global _current_question_id
    _current_question_id = question_id


def get_current_question() -> str:
    """获取当前题目 ID"""
    if _current_question_id is None:
        raise ValueError("未选择题目，请先加载一个题目")
    return _current_question_id


# ============================================================
# Tool 1: query_trains_between_stations
# 对应端点: GET /api/routes
# ============================================================
def _make_func_query_trains_between_stations():
    return {
        "type": "function",
        "function": {
            "name": "query_trains_between_stations",
            "description": "查询从出发站到到达站之间的所有高铁车次，返回车次号及在出发站和到达站的实际到发时间。不返回整趟车的始发站或终到站。",
            "parameters": {
                "type": "object",
                "properties": {
                    "from_station_id": {
                        "type": "string",
                        "description": "出发站电报码，如 VNP（北京南）"
                    },
                    "to_station_id": {
                        "type": "string",
                        "description": "到达站电报码，如 AOH（上海虹桥）"
                    }
                },
                "required": ["from_station_id", "to_station_id"]
            }
        }
    }


# ============================================================
# Tool 2: query_trains_at_station
# 对应端点: GET /api/station/{station_id}
# ============================================================
def _make_func_query_trains_at_station():
    return {
        "type": "function",
        "function": {
            "name": "query_trains_at_station",
            "description": "查询经过指定车站的所有车次，返回车次号及在该站的时间。结果按车次号排序。",
            "parameters": {
                "type": "object",
                "properties": {
                    "station_id": {
                        "type": "string",
                        "description": "车站电报码，如 VNP（北京南）"
                    }
                },
                "required": ["station_id"]
            }
        }
    }


# ============================================================
# Tool 3: query_train_detail
# 对应端点: GET /api/train/{train_num}
# ============================================================
def _make_func_query_train_detail():
    return {
        "type": "function",
        "function": {
            "name": "query_train_detail",
            "description": "查询指定车次的完整经停站详情，返回该车次所有经停站的序号、站名、电报码、时间。",
            "parameters": {
                "type": "object",
                "properties": {
                    "train_num": {
                        "type": "string",
                        "description": "车次号，如 G1"
                    }
                },
                "required": ["train_num"]
            }
        }
    }


# ============================================================
# Tool 4: list_stations
# 对应端点: GET /api/station
# ============================================================
def _make_func_list_stations():
    return {
        "type": "function",
        "function": {
            "name": "list_stations",
            "description": "搜索车站列表，返回车站电报码和中文名。不传 keyword 则返回全部车站（受分页限制）。",
            "parameters": {
                "type": "object",
                "properties": {
                    "keyword": {
                        "type": "string",
                        "description": "搜索关键词，匹配站名或电报码。可选。"
                    },
                    "limit": {
                        "type": "integer",
                        "description": "每页返回数量，默认 50，最大 200"
                    },
                    "page": {
                        "type": "integer",
                        "description": "页码，默认 1"
                    }
                },
                "required": []
            }
        }
    }


# ============================================================
# Tool 5: list_trains
# 对应端点: GET /api/train
# ============================================================
def _make_func_list_trains():
    return {
        "type": "function",
        "function": {
            "name": "list_trains",
            "description": "搜索车次列表，返回所有车次号。不传 keyword 则返回全部车次（受分页限制）。",
            "parameters": {
                "type": "object",
                "properties": {
                    "keyword": {
                        "type": "string",
                        "description": "匹配车次号，如 G1。可选。"
                    },
                    "limit": {
                        "type": "integer",
                        "description": "每页返回数量，默认 50，最大 200"
                    },
                    "page": {
                        "type": "integer",
                        "description": "页码，默认 1"
                    }
                },
                "required": []
            }
        }
    }


# ============================================================
# Tool 6: query_tickets
# 对应端点: GET /api/train/{train_num}/ticket
# ============================================================
def _make_func_query_tickets():
    return {
        "type": "function",
        "function": {
            "name": "query_tickets",
            "description": "查询指定车次各站对之间的余票数量。三种座位类型：class2=二等座（经济型），class1=一等座（更舒适），class0=特等座（可躺，十分舒适）。不指定 seat_types 则返回全部三种。不指定 from_station_id/to_station_id 则返回所有站对。",
            "parameters": {
                "type": "object",
                "properties": {
                    "train_num": {
                        "type": "string",
                        "description": "车次号，如 G1"
                    },
                    "seat_types": {
                        "type": "array",
                        "items": {
                            "type": "string",
                            "enum": ["class0", "class1", "class2"]
                        },
                        "description": "要查询的座位类型列表，不传则返回全部三种"
                    },
                    "from_station_id": {
                        "type": "string",
                        "description": "出发站电报码，不传则返回所有站对"
                    },
                    "to_station_id": {
                        "type": "string",
                        "description": "到达站电报码，不传则返回所有站对"
                    }
                },
                "required": ["train_num"]
            }
        }
    }

# ============================================================
# Tool 7: query_ticket_price
# ============================================================
def _make_func_query_ticket_price():
    return {
        "type": "function",
        "function": {
            "name": "query_ticket_price",
            "description": "查询某车次在指定出发站到到达站之间的各席别票价（元）。返回 class2/class1/class0 的价格。如果查不到该区间的票价数据则返回 null。",
            "parameters": {
                "type": "object",
                "properties": {
                    "train_num": {
                        "type": "string",
                        "description": "车次号，如 G1"
                    },
                    "from_station_id": {
                        "type": "string",
                        "description": "出发站电报码，如 VNP"
                    },
                    "to_station_id": {
                        "type": "string",
                        "description": "到达站电报码，如 JGK"
                    }
                },
                "required": ["train_num", "from_station_id", "to_station_id"]
            }
        }
    }


# ============================================================
# 工具注册列表
# ============================================================
TOOLS = [
    _make_func_query_trains_between_stations(),
    _make_func_query_trains_at_station(),
    _make_func_query_train_detail(),
    _make_func_list_stations(),
    _make_func_list_trains(),
    _make_func_query_tickets(),
    _make_func_query_ticket_price(),
]

# ============================================================
# 工具名称 → handler 映射
# ============================================================
TOOL_HANDLERS = {
    "query_trains_between_stations": None,  # 在 execute_tool_handler 中分发
    "query_trains_at_station": None,
    "query_train_detail": None,
    "list_stations": None,
    "list_trains": None,
    "query_tickets": None,
    "query_ticket_price": None,
}


def get_tool_names() -> List[str]:
    """获取所有工具名称列表"""
    return [t["function"]["name"] for t in TOOLS]

# 预览缓存：存储未确认的自动出题数据（key=question_id）
_preview_cache: Dict[str, Dict] = {}


class _RetryTrainError(HTTPException):
    """内层找不到合法解（中间站/换乘车次/额外站不足等）→ 应换下一辆 T 重试，而非直接报错。"""


def _create_in_memory_question_db() -> sqlite3.Connection:
    """创建内存中的题目数据库（用于预览验证，不做磁盘写入）"""
    conn = sqlite3.connect(':memory:')
    conn.execute("PRAGMA foreign_keys = ON")
    for table in TICKET_TABLES:
        conn.execute(f"""
            CREATE TABLE IF NOT EXISTS {table} (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                train_num TEXT NOT NULL,
                from_station_id TEXT NOT NULL,
                to_station_id TEXT NOT NULL,
                tickets INTEGER NOT NULL CHECK (tickets >= 0 AND tickets <= {QUESTION_CONFIG["ticket_max_value"]}),
                UNIQUE(train_num, from_station_id, to_station_id)
            )
        """)
    return conn
from verifier import verify_final_plan, normalize_final_plan


# ============================================================
# 日志配置
# ============================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger("server")


# ============================================================
# 应用生命周期
# ============================================================
@asynccontextmanager
async def lifespan(app: FastAPI):
    """应用启动和关闭时的生命周期管理"""
    ensure_directories()
    logger.info(f"Benchmark 服务启动 - {API_CONFIG['host']}:{API_CONFIG['port']}")
    yield
    logger.info("Benchmark 服务关闭")


app = FastAPI(title="高铁购票大模型测试平台", version="1.0.0", lifespan=lifespan)

# CORS 配置
app.add_middleware(
    CORSMiddleware,
    allow_origins=API_CONFIG["cors_allow_origins"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ============================================================
# 第一部分：基础数据查询接口
# ============================================================

# --- 1.2 GET /api/train 车次列表 ---
@app.get("/api/train")
def api_list_trains(
    keyword: str = "",
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200)
):
    """返回所有 G 车次号列表，支持关键词搜索和分页"""
    conn = get_railway_conn()
    cursor = conn.cursor()
    try:
        offset = (page - 1) * limit
        if keyword:
            cursor.execute(
                "SELECT COUNT(*) FROM trains WHERE train_num LIKE ?",
                (f"%{keyword}%",)
            )
            total = cursor.fetchone()[0]
            cursor.execute(
                "SELECT train_num FROM trains WHERE train_num LIKE ? ORDER BY train_num LIMIT ? OFFSET ?",
                (f"%{keyword}%", limit, offset)
            )
        else:
            cursor.execute("SELECT COUNT(*) FROM trains")
            total = cursor.fetchone()[0]
            cursor.execute(
                "SELECT train_num FROM trains ORDER BY train_num LIMIT ? OFFSET ?",
                (limit, offset)
            )
        data = [{"train_num": row[0]} for row in cursor.fetchall()]
        total_pages = (total + limit - 1) // limit
        return {
            "total": total,
            "page": page,
            "limit": limit,
            "total_pages": total_pages,
            "data": data,
        }
    finally:
        conn.close()


# --- 1.3 GET /api/station 车站列表 ---
@app.get("/api/station")
def api_list_stations(
    keyword: str = "",
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200)
):
    """返回所有车站列表，支持关键词搜索和分页"""
    conn = get_railway_conn()
    cursor = conn.cursor()
    try:
        offset = (page - 1) * limit
        if keyword:
            cursor.execute(
                "SELECT COUNT(*) FROM stations WHERE station_name LIKE ? OR station_id LIKE ?",
                (f"%{keyword}%", f"%{keyword}%")
            )
            total = cursor.fetchone()[0]
            cursor.execute(
                """SELECT station_id, station_name FROM stations
                   WHERE station_name LIKE ? OR station_id LIKE ?
                   ORDER BY station_name LIMIT ? OFFSET ?""",
                (f"%{keyword}%", f"%{keyword}%", limit, offset)
            )
        else:
            cursor.execute("SELECT COUNT(*) FROM stations")
            total = cursor.fetchone()[0]
            cursor.execute(
                "SELECT station_id, station_name FROM stations ORDER BY station_name LIMIT ? OFFSET ?",
                (limit, offset)
            )
        data = [{"station_id": row[0], "station_name": row[1]} for row in cursor.fetchall()]
        total_pages = (total + limit - 1) // limit
        return {
            "total": total,
            "page": page,
            "limit": limit,
            "total_pages": total_pages,
            "data": data,
        }
    finally:
        conn.close()


# --- GET /api/station/random 随机车站 ---
@app.get("/api/station/random")
def api_station_random():
    """返回一个随机车站"""
    conn = get_railway_conn()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT station_id, station_name FROM stations ORDER BY RANDOM() LIMIT 1")
        row = cursor.fetchone()
        if row is None:
            raise HTTPException(status_code=404, detail="无车站数据")
        return {"station_id": row[0], "station_name": row[1]}
    finally:
        conn.close()


# --- 1.4 GET /api/train/{train_num} 车次详情 ---
@app.get("/api/train/{train_num}")
def api_train_detail(train_num: str):
    """返回指定车次的完整信息（含经停站列表）"""
    conn = get_railway_conn()
    cursor = conn.cursor()
    try:
        # 验证车次是否存在
        cursor.execute("SELECT train_no FROM trains WHERE train_num = ?", (train_num,))
        row = cursor.fetchone()
        if row is None:
            raise HTTPException(status_code=404, detail=f"车次 {train_num} 未找到")
        train_no = row[0]

        # 获取经停站列表
        stops = get_train_stops(conn, train_num)
        return {
            "train_num": train_num,
            "train_no": train_no,
            "stops": stops,
        }
    finally:
        conn.close()


# --- 1.5 GET /api/station/{station_id} 车站详情 ---
@app.get("/api/station/{station_id}")
def api_station_detail(station_id: str):
    """返回指定车站的经停车次列表"""
    conn = get_railway_conn()
    try:
        result = get_station_trains(conn, station_id)
        if result is None:
            raise HTTPException(status_code=404, detail=f"车站 {station_id} 未找到")
        return {
            "station_id": result["station_id"],
            "station_name": result["station_name"],
            "train_count": len(result["data"]),
            "trains": result["data"],
        }
    finally:
        conn.close()


# --- 1.6 GET /api/train/{train_num}/ticket 余票查询 ---
@app.get("/api/train/{train_num}/ticket")
def api_train_ticket(
    train_num: str,
    from_station_id: Optional[str] = Query(None),
    to_station_id: Optional[str] = Query(None),
    seat_types: Optional[str] = Query(None),
    question_id: Optional[str] = Query(None)
):
    """返回指定车次在指定题目下的余票数据"""
    qid = question_id
    if not qid:
        try:
            qid = get_current_question()
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))

    db_path = get_question_db_path(qid)
    if not os.path.exists(db_path):
        raise HTTPException(status_code=404, detail=f"题目 {qid} 不存在")

    seat_type_list = None
    if seat_types:
        seat_type_list = [s.strip() for s in seat_types.split(",") if s.strip() in TICKET_TABLES]

    conn = sqlite3.connect(db_path)
    try:
        tickets = get_train_tickets_filtered(
            conn, train_num, from_station_id, to_station_id, seat_type_list
        )
        return {
            "train_num": train_num,
            "question_id": qid,
            "tickets": tickets,
        }
    finally:
        conn.close()


# --- 1.7 GET /api/routes 两站之间车次 ---
@app.get("/api/routes")
def api_routes(
    from_station_id: str = Query(None),
    to_station_id: str = Query(None)
):
    """查询同时经过出发站和到达站的所有车次"""
    if not from_station_id or not to_station_id:
        raise HTTPException(status_code=400, detail="缺少 from_station_id 或 to_station_id")

    conn = get_railway_conn()
    try:
        data = get_routes_between(conn, from_station_id, to_station_id)
        if not data:
            raise HTTPException(
                status_code=404,
                detail=f"未找到从 {from_station_id} 到 {to_station_id} 的车次"
            )
        return {
            "from_station_id": from_station_id,
            "to_station_id": to_station_id,
            "count": len(data),
            "data": data,
        }
    finally:
        conn.close()


# ============================================================
# 第二部分：出题器相关接口
# ============================================================

# --- 请求模型 ---
class UpdateTicketRequest(BaseModel):
    question_id: str
    train_num: str
    from_station_id: str
    to_station_id: str
    seat_type: str
    tickets: int


class AutoGenerateRequest(BaseModel):
    # question_type：存在性必传；选择性可不传（None）→ 服务端按行为约束自动推导（保证有解）
    question_type: Optional[str] = None  # transfer/short_buy/extra_front/extra_rear/mixed（direct 仅作 mixed 段内策略）
    from_station_id: str  # 可接受站名（如"北京南"）或站ID（如"VNP"），服务端自动解析
    to_station_id: str  # 同上
    mode: str = ""  # "existence" 存在性(0_ 必出 + fake_interference 时另出 1_)；"selective" 选择性(一份 2_)；空串按 selective 处理
    random_tickets: bool = False  # 是否添加干扰票（保留字段，实际以 mode/fake_interference 为准）
    fake_interference: bool = False  # 存在性是否额外生成 1_ 伪干扰（票数 < 人数，仍唯一解）；False=仅出 0_
    interference_density: float = 0.02  # 干扰密度（全局池比例，默认 2%，上限 5%，步长 0.1%）
    transfers: int = 0  # 换乘次数（仅 mixed 题型）
    segment_plans: List[str] = []  # 每段策略，长度 = transfers + 1
    custom_qid: str = ""  # 自定义题名，为空则自动生成
    seed: Optional[int] = None
    people_count: Optional[int] = None  # 需求人数：缺省时服务端随机 3~6
    seat_type: str = "class2"  # 答案票等级（class0/class1/class2）
    # 评判标准（仅选择性题，单选必选；只作为题目对模型的优化要求，无标准答案、不参与核查）：
    #   comprehensive 综合考虑（默认）/ fastest 最快 / cheapest 最便宜 /
    #   depart_latest 最晚出发 / arrive_earliest 最早到达
    criterion: str = "comprehensive"
    # 行为约束（仅选择性题，多选；作为题目对模型的要求，参与硬校验）：
    #   no_transfer 不允许换乘 / no_short_buy_extra 不允许买短补长与额外购买
    constraints: List[str] = []


# --- POST /api/update_ticket 实时更新余票 ---
@app.post("/api/update_ticket")
def api_update_ticket(req: UpdateTicketRequest):
    """实时更新余票数据（手动出题/改题调用）"""
    # 校验座位类型
    if req.seat_type not in TICKET_TABLES:
        raise HTTPException(status_code=400, detail=f"无效的座位类型: {req.seat_type}")
    # 校验余票范围
    if not (0 <= req.tickets <= QUESTION_CONFIG["ticket_max_value"]):
        raise HTTPException(status_code=400, detail=f"余票超出范围: {req.tickets}")

    # 校验车站与车次是否存在
    rw_conn = get_railway_conn()
    try:
        if not validate_train_exists(rw_conn, req.train_num):
            raise HTTPException(status_code=400, detail=f"车次 {req.train_num} 不存在")
        for sid in [req.from_station_id, req.to_station_id]:
            if not validate_station_exists(rw_conn, sid):
                raise HTTPException(status_code=400, detail=f"车站 {sid} 不存在")
    finally:
        rw_conn.close()

    # 更新余票
    db_path = get_question_db_path(req.question_id)
    if not os.path.exists(db_path):
        # 如果题目数据库不存在，自动创建（稀疏存储，不预填充）
        create_question_db(req.question_id)

    q_conn = sqlite3.connect(db_path)
    try:
        db_update_ticket(
            q_conn, req.train_num,
            req.from_station_id, req.to_station_id,
            req.seat_type, req.tickets
        )
        # 更新元数据修改时间
        update_question_metadata(req.question_id, status="draft")
        return {"success": True, "message": "余票已更新"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        q_conn.close()


# ============================================================
# 自动出题器辅助函数
# ============================================================

def _calc_time_diff_minutes(time1: str, time2: str) -> int:
    """计算 time2 - time1 的分钟数（仅当日，不做跨天换算）

    要求同一天内 time2 在 time1 之后：返回正值表示当日间隔；
    若 time2 时钟时间早于 time1（即跨天到次日），返回负值。
    """
    try:
        h1, m1 = map(int, time1.split(":"))
        h2, m2 = map(int, time2.split(":"))
        return (h2 * 60 + m2) - (h1 * 60 + m1)
    except (ValueError, AttributeError):
        return 0


def _get_trains_passing_station(rw_conn: sqlite3.Connection, station_id: str) -> List[str]:
    """获取经过指定车站的所有车次号"""
    cursor = rw_conn.cursor()
    cursor.execute("SELECT DISTINCT train_num FROM train_stops WHERE station_id = ?", (station_id,))
    return [row[0] for row in cursor.fetchall()]


def _train_passes_station(rw_conn: sqlite3.Connection, train_num: str, station_id: str) -> bool:
    """检查车次是否经过指定车站"""
    cursor = rw_conn.cursor()
    cursor.execute("SELECT 1 FROM train_stops WHERE train_num = ? AND station_id = ?", (train_num, station_id))
    return cursor.fetchone() is not None


def _get_routes_for_transfer(rw_conn: sqlite3.Connection,
                              from_id: str, to_id: str) -> List[Dict]:
    """
    获取换乘题型的候选车次 T，要求：
    - T 经过出发站 A
    - T 不经过目标站 B（防止买短补长）
    - T 在 A 之后至少还有一站（有中间站 M 可选）
    """
    all_trains = _get_trains_passing_station(rw_conn, from_id)
    routes = []
    for tn in all_trains:
        if _train_passes_station(rw_conn, tn, to_id):
            continue  # 排除经过 B 的车次
        stops = get_train_stops(rw_conn, tn)
        stop_ids = [s["station_id"] for s in stops]
        if from_id in stop_ids:
            from_idx = stop_ids.index(from_id)
            if from_idx < len(stops) - 1:  # A 不是最后一站
                routes.append({
                    "train_num": tn,
                    "depart_time": stops[from_idx]["stop_time"],
                    "arrive_time": stops[-1]["stop_time"],
                })
    return routes


def _get_city_key(station_name: str) -> str:
    """从站名提取城市标识（用于判断同城车站）"""
    for suf in ['东', '南', '西', '北', '站']:
        if station_name.endswith(suf) and len(station_name) > len(suf) + 1:
            return station_name[:-len(suf)]
    return station_name


def _find_transfer_solutions(rw_conn: sqlite3.Connection,
                             excluded_train_num: str,
                             mid_refs: List[Tuple[str, str]],
                             to_station_id: str,
                             min_gap: int = 20) -> List[Dict]:
    """
    核心一：一次枚举「全部可行换乘方案」(M, U)。

    mid_refs: [(mid_id, 参考到达时间)]，参考时间 = 上一程车到达该 M 的时刻；
              对单一 M 换乘即 [(M, T到达M时间)]。
    对每个 M 找满足以下条件的车 U（一次 SQL，IN 查询全部 M）：
    - U 经过 M 且在 M 之后经过 B（stop_no(M) < stop_no(B)）
    - U != excluded_train_num（不能是同辆车）
    - 参考到达时间 + min_gap <= U 从 M 出发时间
    返回 [{mid_id, train_num, u_m_depart, u_b_arrive, ref_time, gap}]
    """
    if not mid_refs:
        return []
    mid_ids = [m for m, _ in mid_refs]
    ref_map = dict(mid_refs)
    placeholders = ",".join("?" * len(mid_ids))
    cursor = rw_conn.cursor()
    cursor.execute(f"""
        SELECT u_m.station_id, u_m.stop_time, u_b.stop_time, u_m.train_num
        FROM train_stops u_m
        JOIN train_stops u_b
          ON u_b.train_num = u_m.train_num AND u_b.station_id = ?
        WHERE u_m.station_id IN ({placeholders})
          AND u_m.stop_no < u_b.stop_no
          AND u_m.train_num != ?
        ORDER BY u_m.stop_time
    """, [to_station_id] + mid_ids + [excluded_train_num])

    solutions = []
    for mid_id, u_m_depart, u_b_arrive, u_train in cursor.fetchall():
        ref_time = ref_map.get(mid_id)
        if ref_time and u_m_depart:
            gap = _calc_time_diff_minutes(ref_time, u_m_depart)
            if gap >= min_gap:
                solutions.append({
                    "mid_id": mid_id,
                    "train_num": u_train,
                    "u_m_depart": u_m_depart,
                    "u_b_arrive": u_b_arrive,
                    "ref_time": ref_time,
                    "gap": gap,
                })
    return solutions


def _find_transfer_trains(rw_conn: sqlite3.Connection,
                          target_train_num: str,
                          mid_station_id: str,
                          to_station_id: str,
                          mid_stop_time: str,
                          min_gap: int = 20) -> List[Dict]:
    """查找从中间站 M 到 B 的其他车次 U（复用核心一枚举，保持原返回结构）"""
    sols = _find_transfer_solutions(
        rw_conn, target_train_num, [(mid_station_id, mid_stop_time)],
        to_station_id, min_gap,
    )
    return [
        {"train_num": s["train_num"], "mid_depart": s["u_m_depart"], "arrive_time": s["u_b_arrive"]}
        for s in sols
    ]


def _random_solution_tickets(people_count: int) -> int:
    """生成答案票数：保证 ≥ 需求人数（票够），在 1~1.5× 人数范围内随机（而非恒等于人数）。

    例：2 人 → 2~3；4 人 → 4~6；20 人 → 20~30。人数=1 时退化为 1。
    """
    hi = max(people_count, int(people_count * 1.5))
    return random.randint(people_count, hi)





def _build_transfer_segments(q_conn: Optional[sqlite3.Connection],
                             rw_conn: sqlite3.Connection,
                             train_num: str, stops: List[Dict],
                             from_id: str, mid_id: str, u_train_num: str,
                             actual_dest: str, seat_type: str,
                             people_count: int) -> List[Dict]:
    """写换乘题两段：T(A→M) + U(M→B)；q_conn 为 None 时只构建 segments 不写库。

    供出题与「换方案」共用：换方案传 q_conn=None 仅重算预览，确认时再统一写盘。
    """
    station_names = {s["station_id"]: s["station_name"] for s in stops}
    tickets1 = _random_solution_tickets(people_count)
    tickets2 = _random_solution_tickets(people_count)
    if q_conn is not None:
        db_update_ticket(q_conn, train_num, from_id, mid_id, seat_type, tickets1)
        db_update_ticket(q_conn, u_train_num, mid_id, actual_dest, seat_type, tickets2)
    u_stops = get_train_stops(rw_conn, u_train_num)
    u_station_names = {s["station_id"]: s["station_name"] for s in u_stops}
    return [
        {
            "train_num": train_num,
            "from": station_names.get(from_id, from_id),
            "to": station_names.get(mid_id, mid_id),
            "from_station_id": from_id,
            "to_station_id": mid_id,
            "ride_from": station_names.get(from_id, from_id),
            "ride_to": station_names.get(mid_id, mid_id),
            "ride_from_station_id": from_id,
            "ride_to_station_id": mid_id,
            "tickets": tickets1,
            "seat_type": seat_type,
            "seg_type": "purchase",
        },
        {
            "train_num": u_train_num,
            "from": u_station_names.get(mid_id, mid_id),
            "to": u_station_names.get(actual_dest, actual_dest),
            "from_station_id": mid_id,
            "to_station_id": actual_dest,
            "ride_from": u_station_names.get(mid_id, mid_id),
            "ride_to": u_station_names.get(actual_dest, actual_dest),
            "ride_from_station_id": mid_id,
            "ride_to_station_id": actual_dest,
            "tickets": tickets2,
            "seat_type": seat_type,
            "seg_type": "purchase",
        },
    ]


def _write_legal_solution(q_conn: sqlite3.Connection,
                          rw_conn: sqlite3.Connection,
                          question_type: str,
                          train_num: str,
                          stops: List[Dict],
                          from_id: str, to_id: str,
                          from_idx: int, to_idx: int,
                          segment_plans: List[str] = None,
                          transfer_dest: str = None,
                          seat_type: str = "class2",
                          people_count: int = 1,
                          min_gap: int = 20) -> Dict:
    """
    写入合法解（正数票）。
    返回包含 segments 和 target_mid_idx 的字典。
    """
    segments = []
    target_mid_idx = None  # 记录 short_buy 的目标中间站索引
    solutions = None       # 核心一：本车次全部可行换乘方案（仅 transfer）
    chosen_solution = None # 当前选中的方案（换方案时使用）

    # 获取站名映射
    station_names = {s["station_id"]: s["station_name"] for s in stops}

    if question_type == "transfer":
        middle_indices = [i for i in range(from_idx + 1, to_idx)]

        # 过滤掉与目标站同城的中转站
        if transfer_dest:
            cur = rw_conn.cursor()
            cur.execute("SELECT station_name FROM stations WHERE station_id = ?", (transfer_dest,))
            row = cur.fetchone()
            dest_station_name = row[0] if row else transfer_dest
            dest_city = _get_city_key(dest_station_name)
            middle_indices = [
                i for i in middle_indices
                if _get_city_key(station_names.get(stops[i]["station_id"], "")) != dest_city
            ]

        if not middle_indices:
            raise _RetryTrainError(status_code=400, detail="目标车次无合适的中间站")

        # 实际目标站：用户输入的真正目的地，而非 T 的末站
        actual_dest = transfer_dest if transfer_dest else to_id

        # 核心一：一次枚举本车次全部可行 (M, U)，均匀随机挑一个
        mid_refs = [(stops[i]["station_id"], stops[i]["stop_time"]) for i in middle_indices]
        solutions = _find_transfer_solutions(rw_conn, train_num, mid_refs, actual_dest, min_gap)
        if not solutions:
            raise _RetryTrainError(
                status_code=400,
                detail=f"未找到合适的换乘车次到 {actual_dest}（已枚举全部中间站）"
            )
        chosen_solution = random.choice(solutions)
        mid_id = chosen_solution["mid_id"]
        mid_stop_time = chosen_solution["ref_time"]
        u_train = chosen_solution  # dict 含 train_num
        target_mid_idx = next(i for i in middle_indices if stops[i]["station_id"] == mid_id)

        segments = _build_transfer_segments(
            q_conn, rw_conn, train_num, stops, from_id, mid_id,
            u_train["train_num"], actual_dest, seat_type, people_count,
        )

    elif question_type == "short_buy":
        middle_indices = [i for i in range(from_idx + 1, to_idx)]
        if not middle_indices:
            raise _RetryTrainError(status_code=400, detail="目标车次无合适的中间站")
        target_mid_idx = random.choice(middle_indices)
        mid_id = stops[target_mid_idx]["station_id"]
        tickets = _random_solution_tickets(people_count)
        db_update_ticket(q_conn, train_num, from_id, mid_id, seat_type, tickets)
        # 买短补长：买 A→M 有票，实际乘坐 A→B（M→B 无票，上车补票）
        segments.append({
            "train_num": train_num,
            "from": station_names.get(from_id, from_id),
            "to": station_names.get(mid_id, mid_id),
            "from_station_id": from_id,
            "to_station_id": mid_id,
            "ride_from": station_names.get(from_id, from_id),
            "ride_to": station_names.get(to_id, to_id),
            "ride_from_station_id": from_id,
            "ride_to_station_id": to_id,
            "tickets": tickets,
            "seat_type": seat_type,
            "seg_type": "purchase",
        })

    elif question_type == "extra_front":
        # 前额外：B→C 没票，往前找 k 站 (1~3) 买 Aₖ→C
        max_extra = min(from_idx, 3)
        if max_extra < 1:
            raise _RetryTrainError(status_code=400, detail="出发站前无足够车站做前额外")
        k = random.randint(1, max_extra)
        extra_from_idx = from_idx - k
        extra_from_id = stops[extra_from_idx]["station_id"]
        tickets = _random_solution_tickets(people_count)
        db_update_ticket(q_conn, train_num, extra_from_id, to_id, seat_type, tickets)
        segments.append({
            "train_num": train_num,
            "from": station_names.get(extra_from_id, extra_from_id),
            "to": station_names.get(to_id, to_id),
            "from_station_id": extra_from_id,
            "to_station_id": to_id,
            "ride_from": station_names.get(from_id, from_id),
            "ride_to": station_names.get(to_id, to_id),
            "ride_from_station_id": from_id,
            "ride_to_station_id": to_id,
            "tickets": tickets,
            "seat_type": seat_type,
            "seg_type": "purchase",
        })

    elif question_type == "extra_rear":
        # 后额外：B→C 没票，往后找 k 站 (1~3) 买 B→Dₖ
        max_extra = min(len(stops) - 1 - to_idx, 3)
        if max_extra < 1:
            raise _RetryTrainError(status_code=400, detail="到达站后无足够车站做后额外")
        k = random.randint(1, max_extra)
        extra_to_idx = to_idx + k
        extra_to_id = stops[extra_to_idx]["station_id"]
        tickets = _random_solution_tickets(people_count)
        db_update_ticket(q_conn, train_num, from_id, extra_to_id, seat_type, tickets)
        segments.append({
            "train_num": train_num,
            "from": station_names.get(from_id, from_id),
            "to": station_names.get(extra_to_id, extra_to_id),
            "from_station_id": from_id,
            "to_station_id": extra_to_id,
            "ride_from": station_names.get(from_id, from_id),
            "ride_to": station_names.get(to_id, to_id),
            "ride_from_station_id": from_id,
            "ride_to_station_id": to_id,
            "tickets": tickets,
            "seat_type": seat_type,
            "seg_type": "purchase",
        })

    elif question_type == "mixed" and segment_plans:
        # 混合：设换乘数 N = len(segment_plans) - 1
        # 路径：A → M1 → M2 → ... → MN → B
        # 段 0 用 T (A→M1)，段 i>0 在 Mi 换乘后到下一站
        transfers = len(segment_plans) - 1

        # 选取 N 个中间站
        middle_indices = [i for i in range(from_idx + 1, to_idx)]
        if len(middle_indices) < transfers:
            raise _RetryTrainError(
                status_code=400,
                detail=f"混合题型需要至少 {transfers} 个中间站，当前只有 {len(middle_indices)}"
            )
        selected = sorted(random.sample(middle_indices, transfers)) if transfers > 0 else []
        mid_ids = [stops[i]["station_id"] for i in selected]
        path_nodes = [from_id] + mid_ids + [to_id]  # 长度 = transfers + 2

        current_train = train_num
        current_stops = stops
        # 乘客到达当前段起点站的时间（换乘衔接基准，跨段跟踪实际到达时间）
        arrival_at_seg_start = None
        stop_ids_for_t = [s["station_id"] for s in stops]

        for seg_idx, strategy in enumerate(segment_plans):
            seg_from = path_nodes[seg_idx]
            seg_to = path_nodes[seg_idx + 1]

            # 段 > 0：需要找换乘车次，衔接基准 = 上一段车次到达本段起点的时间
            if seg_idx > 0:
                candidates = _find_transfer_trains(rw_conn, train_num, seg_from, seg_to, arrival_at_seg_start, min_gap)
                # 非最后一段的买短补长：排除经过最终站 B 的车次，避免模型跳过后续换乘直达终点
                if strategy == "short_buy" and seg_idx < transfers:
                    candidates = [
                        c for c in candidates
                        if not _train_passes_station(rw_conn, c["train_num"], to_id)
                    ]
                    if not candidates:
                        raise _RetryTrainError(
                            status_code=400,
                            detail=f"段 {seg_idx} 买短补长无可用的换乘车次（候选均经过终点站）"
                        )
                if not candidates:
                    raise _RetryTrainError(status_code=400, detail=f"未找到从 {seg_from} 到 {seg_to} 的换乘车次")
                if seg_idx < transfers:
                    # 后续还有换乘：必须确保选中车次在该段终点（下一换乘点）有有效到达时间
                    valid_candidates = [
                        c for c in candidates
                        if re.match(r'^\d{2}:\d{2}$', c.get("arrive_time") or "")
                    ]
                    if not valid_candidates:
                        raise _RetryTrainError(status_code=400, detail=f"未找到有有效到达时间的换乘车次从 {seg_from} 到 {seg_to}")
                    chosen = random.choice(valid_candidates)
                else:
                    chosen = random.choice(candidates)
                current_train = chosen["train_num"]
                current_stops = get_train_stops(rw_conn, current_train)
                # 更新基准时间：本段车次到达本段终点的时间（下一换乘点）
                arrival_at_seg_start = chosen["arrive_time"]
            else:
                # 段 0 用目标车次 T：记录 T 到达第一换乘点的时间作为衔接基准
                seg_to_idx = stop_ids_for_t.index(seg_to)
                arrival_at_seg_start = stops[seg_to_idx]["stop_time"]

            # 在当前车次上应用段策略
            current_ids = [s["station_id"] for s in current_stops]
            current_names = {s["station_id"]: s["station_name"] for s in current_stops}
            try:
                cur_from_idx = current_ids.index(seg_from)
                cur_to_idx = current_ids.index(seg_to)
            except ValueError:
                raise _RetryTrainError(status_code=400, detail=f"车次 {current_train} 不经过 {seg_from} 或 {seg_to}")

            if strategy == "direct":
                tickets = _random_solution_tickets(people_count)
                db_update_ticket(q_conn, current_train, seg_from, seg_to, seat_type, tickets)
                segments.append({
                    "train_num": current_train,
                    "from": current_names.get(seg_from, seg_from),
                    "to": current_names.get(seg_to, seg_to),
                    "from_station_id": seg_from,
                    "to_station_id": seg_to,
                    "ride_from": current_names.get(seg_from, seg_from),
                    "ride_to": current_names.get(seg_to, seg_to),
                    "ride_from_station_id": seg_from,
                    "ride_to_station_id": seg_to,
                    "tickets": tickets,
                    "seat_type": seat_type,
                    "strategy": strategy,
                    "actual_from": current_names.get(seg_from, seg_from),
                    "actual_to": current_names.get(seg_to, seg_to),
                    "seg_type": "purchase",
                })
            elif strategy == "short_buy":
                mid_indices = [i for i in range(cur_from_idx + 1, cur_to_idx)]
                if not mid_indices:
                    raise _RetryTrainError(status_code=400, detail=f"段 {seg_idx} 买短补长无中间站")
                mid_idx = random.choice(mid_indices)
                mid_id = current_stops[mid_idx]["station_id"]
                tickets = _random_solution_tickets(people_count)
                db_update_ticket(q_conn, current_train, seg_from, mid_id, seat_type, tickets)
                # 买短补长：买 seg_from→mid_id 有票，实际乘坐 seg_from→seg_to（mid→seg_to 补票）
                segments.append({
                    "train_num": current_train,
                    "from": current_names.get(seg_from, seg_from),
                    "to": current_names.get(mid_id, mid_id),
                    "from_station_id": seg_from,
                    "to_station_id": mid_id,
                    "ride_from": current_names.get(seg_from, seg_from),
                    "ride_to": current_names.get(seg_to, seg_to),
                    "ride_from_station_id": seg_from,
                    "ride_to_station_id": seg_to,
                    "tickets": tickets,
                    "seat_type": seat_type,
                    "strategy": strategy,
                    "actual_from": current_names.get(seg_from, seg_from),
                    "actual_to": current_names.get(seg_to, seg_to),
                    "seg_type": "purchase",
                })
            elif strategy == "extra_front":
                max_extra = min(cur_from_idx, 3)
                if max_extra < 1:
                    raise _RetryTrainError(status_code=400, detail=f"段 {seg_idx} 前额外无足够车站")
                k = random.randint(1, max_extra)
                extra_from_id = current_stops[cur_from_idx - k]["station_id"]
                tickets = _random_solution_tickets(people_count)
                db_update_ticket(q_conn, current_train, extra_from_id, seg_to, seat_type, tickets)
                segments.append({
                    "train_num": current_train,
                    "from": current_names.get(extra_from_id, extra_from_id),
                    "to": current_names.get(seg_to, seg_to),
                    "from_station_id": extra_from_id,
                    "to_station_id": seg_to,
                    "ride_from": current_names.get(seg_from, seg_from),
                    "ride_to": current_names.get(seg_to, seg_to),
                    "ride_from_station_id": seg_from,
                    "ride_to_station_id": seg_to,
                    "tickets": tickets,
                    "seat_type": seat_type,
                    "strategy": strategy,
                    "actual_from": current_names.get(seg_from, seg_from),
                    "actual_to": current_names.get(seg_to, seg_to),
                    "seg_type": "purchase",
                })
            elif strategy == "extra_rear":
                max_extra = min(len(current_stops) - 1 - cur_to_idx, 3)
                if max_extra < 1:
                    raise _RetryTrainError(status_code=400, detail=f"段 {seg_idx} 后额外无足够车站")
                k = random.randint(1, max_extra)
                extra_to_id = current_stops[cur_to_idx + k]["station_id"]
                tickets = _random_solution_tickets(people_count)
                db_update_ticket(q_conn, current_train, seg_from, extra_to_id, seat_type, tickets)
                segments.append({
                    "train_num": current_train,
                    "from": current_names.get(seg_from, seg_from),
                    "to": current_names.get(extra_to_id, extra_to_id),
                    "from_station_id": seg_from,
                    "to_station_id": extra_to_id,
                    "ride_from": current_names.get(seg_from, seg_from),
                    "ride_to": current_names.get(seg_to, seg_to),
                    "ride_from_station_id": seg_from,
                    "ride_to_station_id": seg_to,
                    "tickets": tickets,
                    "seat_type": seat_type,
                    "strategy": strategy,
                    "actual_from": current_names.get(seg_from, seg_from),
                    "actual_to": current_names.get(seg_to, seg_to),
                    "seg_type": "purchase",
                })

    return {
        "segments": segments,
        "target_mid_idx": target_mid_idx,
        "solutions": solutions,
        "chosen_solution": chosen_solution,
    }


def _clear_direct_route(q_conn: sqlite3.Connection, train_num: str,
                         from_id: str, to_id: str):
    """删除用户声称区间 (B→C) 的所有余票，确保模型无法直达"""
    for table in TICKET_TABLES:
        q_conn.execute(f"""
            DELETE FROM {table}
            WHERE train_num = ? AND from_station_id = ? AND to_station_id = ?
        """, (train_num, from_id, to_id))
    q_conn.commit()


# 经停站全量缓存（一次性加载，供出题/干扰注入复用，避免逐车次查询）
_STOPS_CACHE: Optional[Dict[str, List[Dict]]] = None


def _get_all_stops_cached(rw_conn: sqlite3.Connection) -> Dict[str, List[Dict]]:
    """一次性加载全部车次的经停站（按 stop_no 排序），之后复用。

    全库约 6 万条经停记录，首次加载约 0.8s，此后每次出题直接在内存构建候选池（毫秒级），
    替代原先逐车次查询（每次出题遍历 6000+ 车次各查一次 DB）。
    """
    global _STOPS_CACHE
    if _STOPS_CACHE is None:
        _STOPS_CACHE = {}
        cur = rw_conn.cursor()
        cur.execute(
            "SELECT train_num, stop_no, station_id, station_name, stop_time "
            "FROM train_stops ORDER BY train_num, stop_no"
        )
        for row in cur.fetchall():
            _STOPS_CACHE.setdefault(row[0], []).append({
                "stop_no": row[1],
                "station_id": row[2],
                "station_name": row[3],
                "stop_time": row[4],
            })
    return _STOPS_CACHE


def _add_interference_all_trains(q_conn: sqlite3.Connection,
                                 rw_conn: sqlite3.Connection,
                                 target_train_num: str,
                                 target_solution_pairs: set,
                                 density: float = 0.02,
                                 block_pairs: set = None,
                                 fake: bool = True,
                                 people_count: int = 2,
                                 shortbuy_guard_from: str = None,
                                 shortbuy_guard_to: str = None):
    """全局池注入干扰（统一密度语义）。

    候选池 = 全部车次的全部 i<j 站对（跳过合法解站对 / block_pairs / 买短补长逃逸对），
    一次 shuffle 后按密度取前 int(池大小 × density) 个 —— 分母固定为全局池大小（约 46 万），
    不再逐车取整导致低密度归零，且跨题可比。

    - block_pairs: 全局禁止站对（如真干扰的直达 (A,B)），所有车次一律跳过；
      伪干扰（fake=True）票数严格 < 人数、无逃逸可能，传空集即可。
    - fake=True : 伪干扰票数 randint(1, 人数-1)，严格不足 → 唯一解（人数=1 时退化为无干扰）
    - fake=False: 真干扰票数 randint(0.5×人数, 1.5×人数)，部分够票（真替代）、部分差一点（陷阱）
    - shortbuy_guard_from/to: 非空时（真干扰的 transfer/mixed），对所有“先经过 A 再经过 B”的车次，
      跳过 (A, X)（X 在 B 之前）站对 —— 防止“买短补长直达 B”的逃逸。
    """
    if block_pairs is None:
        block_pairs = set()
    # 伪干扰边界：1 人时伪干扰票数需 <1（=0 张即无票），退化为无干扰，保证唯一解
    if fake and people_count <= 1:
        return

    stops_cache = _get_all_stops_cached(rw_conn)
    pool: List[tuple] = []
    for train_num, stops in stops_cache.items():
        if len(stops) < 2:
            continue
        ids = [s["station_id"] for s in stops]
        n = len(ids)
        # 买短补长逃逸防护：该车是否“先经过 A 再经过 B”
        guard = None
        if shortbuy_guard_from and shortbuy_guard_to \
                and shortbuy_guard_from in ids and shortbuy_guard_to in ids:
            i_a, i_b = ids.index(shortbuy_guard_from), ids.index(shortbuy_guard_to)
            if i_a < i_b:
                guard = (i_a, i_b)
        for i in range(n):
            for j in range(i + 1, n):
                f, t = ids[i], ids[j]
                if train_num == target_train_num and (f, t) in target_solution_pairs:
                    continue
                if (f, t) in block_pairs:
                    continue
                if guard is not None and i == guard[0] and j < guard[1]:
                    continue
                pool.append((train_num, f, t))

    random.shuffle(pool)
    selected = pool[:int(len(pool) * density)]

    for train_num, f, t in selected:
        seat = random.choice(TICKET_TABLES)
        if fake:
            tickets = random.randint(1, max(1, people_count - 1))  # 伪干扰：票数 < 人数（严格不足，唯一解）
        else:
            lo = max(1, int(people_count * 0.5))
            hi = max(lo, int(people_count * 1.5))
            tickets = random.randint(lo, hi)  # 真干扰：0.5×人数 ~ 1.5×人数
        q_conn.execute(f"""
            INSERT OR REPLACE INTO {seat}
            (train_num, from_station_id, to_station_id, tickets)
            VALUES (?, ?, ?, ?)
        """, (train_num, f, t, tickets))
    q_conn.commit()


# --- POST /api/auto_generate auto出题器生成题目 ---
# 生成单份题的预览（内存 DB 验证，不写磁盘）。
# 存在性出两份：0_无伪干扰 / 1_有伪干扰；选择性出一份：2_。前缀自动加在题名前，与输入无关。
def _generate_one_variant(req, prefix, fake, mode, rw_conn, with_interference, base_seed=None):
    """生成一份题的预览。
    prefix: 题名前缀（存在性 "0_"/"1_"；选择性 "2_"）
    fake: 伪干扰模式（干扰票数严格 < 人数，保证唯一解）
    mode: 题目模式（"existence"/"selective"，用于 type 显示）
    with_interference: 是否添加干扰票（0_=False 完全无干扰，唯一解）
    base_seed: 共用随机种子（存在性两份共用 → 同一车次同一合法解）
    返回 (question_id, preview_dict)；重试耗尽抛 HTTPException(400)
    """
    # 选择性题题型由行为约束自动推导（手动/批量一致，保证有解）：
    #   constraints 含 no_transfer → 买短补长（换乘被禁时唯一可达策略）
    #   否则（无约束或仅 no_short_buy_extra）→ 换乘（一次换乘保证可达）
    if mode == "selective":
        req.question_type = "short_buy" if "no_transfer" in (req.constraints or []) else "transfer"
    elif req.question_type is None:
        raise HTTPException(status_code=400, detail="存在性题必须指定题型")

    valid_types = ["transfer", "short_buy", "extra_front", "extra_rear", "mixed"]
    if req.question_type not in valid_types:
        raise HTTPException(status_code=400, detail=f"无效的题型: {req.question_type}")

    # 需求人数：缺省时服务端随机 3~6（手动页每次进入/重新出题时前端也会随机，双保险）
    if req.people_count is None:
        req.people_count = random.randint(3, 6)
    if not (1 <= req.people_count <= QUESTION_CONFIG["max_people_count"]):
        raise HTTPException(status_code=400, detail=f"人数超出范围: {req.people_count}")
    if req.seat_type not in TICKET_TABLES:
        raise HTTPException(status_code=400, detail=f"无效的座位类型: {req.seat_type}")

    if req.question_type == "mixed":
        if req.transfers < 1:
            raise HTTPException(status_code=400, detail="混合题型换乘数至少为 1")
        if len(req.segment_plans) != req.transfers + 1:
            raise HTTPException(status_code=400, detail=f"段策略数需等于换乘数+1（{req.transfers + 1}）")

    # 行为约束（仅选择性题）：白名单校验（不限制题型选择，约束作为对模型输出的硬要求参与核查）
    VALID_CONSTRAINTS = {"no_transfer", "no_short_buy_extra"}
    constraints = req.constraints or []
    invalid = [c for c in constraints if c not in VALID_CONSTRAINTS]
    if invalid:
        raise HTTPException(status_code=400, detail=f"无效的行为约束: {invalid}")
    if constraints and mode != "selective":
        raise HTTPException(status_code=400, detail="行为约束仅选择性题支持")
    # 评判标准（仅选择性题，单选必选）：枚举校验
    VALID_CRITERIA = {"comprehensive", "fastest", "cheapest", "depart_latest", "arrive_earliest"}
    if req.criterion not in VALID_CRITERIA:
        raise HTTPException(status_code=400, detail=f"无效的评判标准: {req.criterion}")
    if mode != "selective" and req.criterion != "comprehensive":
        raise HTTPException(status_code=400, detail="评判标准仅选择性题支持")

    # 换乘衔接最短分钟：所有题型固定 20 分钟（时间约束配置已移除）
    min_gap = 20

    # Step 0: 固定随机种子
    if base_seed is not None:
        random.seed(base_seed)

    # Step 1: 解析车站
    from_id = resolve_station_name_or_id(rw_conn, req.from_station_id)
    to_id = resolve_station_name_or_id(rw_conn, req.to_station_id)

    if not validate_station_exists(rw_conn, from_id):
        raise HTTPException(status_code=400, detail=f"出发站 {from_id} 不存在")
    if not validate_station_exists(rw_conn, to_id):
        raise HTTPException(status_code=400, detail=f"到达站 {to_id} 不存在")

    # Step 2: 获取所有途经车次，随机打乱以便重试
    # 换乘/混合题型：T 不能经过目标站 B（防止 T 直达 B 逃逸）
    if req.question_type in ("transfer", "mixed"):
        routes = _get_routes_for_transfer(rw_conn, from_id, to_id)
    else:
        routes = get_routes_between(rw_conn, from_id, to_id)
    if not routes:
        raise HTTPException(status_code=400, detail=f"未找到从 {from_id} 到 {to_id} 的车次")

    random.shuffle(routes)
    max_attempts = min(len(routes), 15)

    # 题号查重基准：读取一次 metadata（确认时会落盘到同一份，避免覆盖同名题）
    existing_metadata = load_metadata()

    last_error = None
    for attempt in range(max_attempts):
        target_train_num = routes[attempt]["train_num"]

        stops = get_train_stops(rw_conn, target_train_num)
        stop_ids_list = [s["station_id"] for s in stops]
        from_idx = stop_ids_list.index(from_id)
        if req.question_type == "transfer":
            # 换乘：T 不经过 B，用最后一站作为 to_idx
            to_idx = len(stops) - 1
            cur_from_id, cur_to_id = from_id, stops[-1]["station_id"]
            cur_from_idx, cur_to_idx = from_idx, to_idx
        elif req.question_type == "mixed":
            # 混合：T 不经过 B（防止 T 直达 B 逃逸），用最后一站作为 to_idx；
            # 实际终点仍是用户输入的 B，由最后一段换乘车次到达
            to_idx = len(stops) - 1
            cur_from_id, cur_to_id = from_id, to_id
            cur_from_idx, cur_to_idx = from_idx, to_idx
        else:
            to_idx = stop_ids_list.index(to_id)
            if from_idx > to_idx:
                cur_from_id, cur_to_id = to_id, from_id
                cur_from_idx, cur_to_idx = to_idx, from_idx
            else:
                cur_from_id, cur_to_id = from_id, to_id
                cur_from_idx, cur_to_idx = from_idx, to_idx

        # 预检中间站数量
        middle_count = cur_to_idx - cur_from_idx - 1
        if req.question_type not in ("extra_front", "extra_rear") and middle_count < 1:
            continue
        if req.question_type == "mixed":
            required_mid = req.transfers  # 需要的中间站数 = 换乘数
            if middle_count < required_mid:
                last_error = f"中间站不足：需要 {required_mid} 个，最多 {middle_count} 个"
                continue

        # extra 题型要求前后有足够站
        if req.question_type == "extra_front" and cur_from_idx < 1:
            last_error = "出发站前无足够车站做前额外"
            continue
        if req.question_type == "extra_rear" and cur_to_idx >= len(stops) - 1:
            last_error = "到达站后无足够车站做后额外"
            continue

        # Step 3: 确定题名（前缀自动加，与输入无关；仅用于预览标识，不创建磁盘文件）
        if req.custom_qid:
            question_id = prefix + req.custom_qid
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            question_id = f"{prefix}{timestamp}_{attempt}"

        # 题号查重：db 文件 / metadata / 预览缓存 任一命中即视为已占用
        # （custom_qid 与自动题号都查；自动题号同秒撞号时换下一 attempt 重试）
        if (os.path.exists(get_question_db_path(question_id))
                or question_id in _preview_cache
                or question_id in existing_metadata):
            last_error = f"题名 {question_id} 已存在"
            continue

        # 使用内存数据库做验证，不写入磁盘
        q_conn = _create_in_memory_question_db()

        try:
            # Step 4: 写入合法解（写入内存 DB，仅做验证）
            transfer_dest = None
            if req.question_type == "transfer":
                transfer_dest = to_id  # 用户输入的真正目的地
            solution_result = _write_legal_solution(
                q_conn, rw_conn,
                question_type=req.question_type,
                train_num=target_train_num,
                stops=stops,
                from_id=cur_from_id, to_id=cur_to_id,
                from_idx=cur_from_idx, to_idx=cur_to_idx,
                segment_plans=req.segment_plans,
                transfer_dest=transfer_dest,
                seat_type=req.seat_type,
                people_count=req.people_count,
                min_gap=min_gap,
            )
            solution_segments = solution_result["segments"]

            # 收集合法解占用的站对（使用站 ID）
            solution_pairs = set()
            for seg in solution_segments:
                solution_pairs.add((seg["from_station_id"], seg["to_station_id"]))

            # 验证干扰票写入（仍在内存 DB 中）：0_ 完全无干扰；1_/2_ 按 with_interference 添加
            if with_interference:
                # 1_ 伪干扰：票数严格 < 人数，无逃逸可能，不设 block、不清直达
                if fake:
                    block_pairs: set = set()
                    guard_from = guard_to = None
                else:
                    # 2_ 真干扰：全局禁直达 (A,B)；transfer/mixed 额外防“买短补长到 B”
                    block_pairs = {(from_id, to_id)}
                    guard_from, guard_to = (
                        (from_id, to_id) if req.question_type in ("transfer", "mixed")
                        else (None, None)
                    )
                _add_interference_all_trains(
                    q_conn, rw_conn, target_train_num,
                    solution_pairs, req.interference_density,
                    block_pairs=block_pairs, fake=fake, people_count=req.people_count,
                    shortbuy_guard_from=guard_from, shortbuy_guard_to=guard_to,
                )
                if not fake:
                    _clear_direct_route(q_conn, target_train_num, from_id, to_id)

            q_conn.commit()

            # 构造预览
            station_names = {s["station_id"]: s["station_name"] for s in stops}
            from_name = station_names.get(from_id, from_id)
            cursor = rw_conn.cursor()
            cursor.execute("SELECT station_name FROM stations WHERE station_id = ?", (to_id,))
            row = cursor.fetchone()
            dest_name = row[0] if row else to_id
            first_stop = stops[0]["station_name"] if stops else ""
            last_stop = stops[-1]["station_name"] if stops else ""
            # 题目描述：用户填写的初始站到终点站
            question_str = f"{from_name}到{dest_name}"

            if req.question_type == "transfer":
                parts = [f"乘坐 {target_train_num} 从 {from_name} 到 {solution_segments[0]['to']}（有票）"]
                parts.append(f"换乘 {solution_segments[1]['train_num']} 从 {solution_segments[1]['from']} 到 {dest_name}（有票）")
                path_desc = "，".join(parts)
            elif req.question_type == "short_buy":
                path_desc = f"乘坐 {target_train_num} 从 {from_name} 到 {solution_segments[0]['to']}（有票），补票段 {solution_segments[0]['to']}→{dest_name} 无余票需上车补票"
            elif req.question_type == "extra_front":
                path_desc = f"乘坐 {target_train_num} 从 {solution_segments[0]['from']} 到 {dest_name} 有票（前额外），可在 {from_name} 站上车，实际乘坐 {from_name}→{dest_name}"
            elif req.question_type == "extra_rear":
                path_desc = f"乘坐 {target_train_num} 从 {from_name} 到 {solution_segments[0]['to']} 有票（后额外），在 {dest_name} 站提前下车，实际乘坐 {from_name}→{dest_name}"
            elif req.question_type == "mixed":
                # 各分段叙述叠加（与换方案共用 _build_mixed_path_desc）
                path_desc = _build_mixed_path_desc(solution_segments)
            else:
                path_desc = ""

            # 缓存生成的完整数据（含站 ID，供确认时写入磁盘）
            _preview_cache[question_id] = {
                "question_type": req.question_type,
                "segments": solution_segments,
                "target_train_num": target_train_num,
                "cur_from_id": cur_from_id,
                "cur_to_id": cur_to_id,
                "start_station_id": from_id,
                "end_station_id": to_id,
                "stops": stops,
                "random_tickets": with_interference,
                "fake_interference": fake,
                "question_mode": mode,
                "interference_density": req.interference_density,
                "segment_plans": req.segment_plans,
                "interference": with_interference,
                "people_count": req.people_count,
                "seat_type": req.seat_type,
                "criterion": req.criterion if mode == "selective" else "comprehensive",
                "constraints": constraints,  # 行为约束（仅选择性）
                "question": question_str,
                "solutions": solution_result.get("solutions"),
                "chosen_solution": solution_result.get("chosen_solution"),
            }

            q_conn.close()

            return question_id, {
                "question": question_str,
                "question_type": req.question_type,
                "target_train_num": target_train_num,
                "target_train_route": f"{first_stop}→{last_stop}",
                "target_section": f"{from_name}→{dest_name}",
                "path_description": path_desc,
                "solution_segments": solution_segments,
                "criterion": req.criterion if mode == "selective" else "comprehensive",
                "constraints": constraints,
            }

        except HTTPException as e:
            q_conn.close()
            detail = e.detail if hasattr(e, 'detail') else str(e)
            # 内层找不到合法解（_RetryTrainError）或命中重试关键词 → 换下一辆 T 重试，不直接报错
            if isinstance(e, _RetryTrainError) or any(keyword in str(detail) for keyword in [
                "目标车次无合适的中间站",
                "未找到合适的换乘车次",
                "混合策略换乘",
                "出发站前无足够车站",
                "到达站后无足够车站",
            ]):
                last_error = str(detail)
                continue  # 重试下一个车次
            raise  # 非重试性错误，直接抛出

        except Exception as e:
            q_conn.close()
            last_error = f"出题过程异常: {str(e)}"
            continue

    # 所有重试都失败
    raise HTTPException(
        status_code=400,
        detail=f"所有车次重试后仍无法生成题目: {last_error or '未知错误'}"
    )


@app.post("/api/auto_generate")
def api_auto_generate(req: AutoGenerateRequest):
    """自动生成题目。
    - 存在性（mode=existence）：必出 0_（无干扰，唯一解，对标标答）；
      若 fake_interference=true 额外出 1_（伪干扰，票数严格 < 人数，仍唯一解）
    - 选择性（mode=selective）：出一份 2_（真干扰 0.5~1.5×人数）
    - 前缀自动加在题名前，与用户输入无关
    """
    # 生成新题前，清除所有未确认的旧预览缓存（未保存的题自动作废）
    _preview_cache.clear()

    # 决定生成哪些变体：(前缀, 伪干扰, 是否加干扰票)
    # 0_=完全无干扰（唯一解，对标标答）；1_=伪干扰（票数<人数，唯一解）；2_=真干扰
    if req.mode == "existence":
        variants = [("0_", False, False)]
        if req.fake_interference:
            variants.append(("1_", True, True))
    else:
        variants = [("2_", False, True)]

    # 存在性两份共用同一种子 → 同一车次同一合法解，仅干扰不同；选择性单份用 req.seed
    if len(variants) > 1:
        base_seed = req.seed if req.seed is not None else random.randint(0, 2 ** 31)
    else:
        base_seed = req.seed

    rw_conn = get_railway_conn()
    try:
        questions = []
        for prefix, fake, with_if in variants:
            mode = "existence" if prefix != "2_" else "selective"
            question_id, preview = _generate_one_variant(
                req, prefix, fake, mode, rw_conn, with_if, base_seed=base_seed,
            )
            questions.append({"question_id": question_id, "preview": preview})

        return {
            "success": True,
            "questions": questions,
            "message": f"生成 {len(questions)} 道题（预览，尚未保存）",
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        rw_conn.close()


# --- GET /api/auto_generate/previews 查看预览缓存（调试用） ---
@app.get("/api/auto_generate/previews")
def api_auto_generate_previews():
    """查看当前内存中的预览缓存"""
    return {
        "count": len(_preview_cache),
        "previews": [
            {
                "question_id": qid,
                "question_type": data.get("question_type"),
                "question": data.get("question"),
                "target_train_num": data.get("target_train_num"),
                "segments_count": len(data.get("segments", [])),
                "segment_plans": data.get("segment_plans"),
            }
            for qid, data in _preview_cache.items()
        ]
    }


class ConfirmAutoGenerateRequest(BaseModel):
    question_id: str
    question_type: str = ""
    answer: str = ""
    interference: bool = False
    interference_density: float = 0.0


class ClearPreviewRequest(BaseModel):
    question_id: str


# --- POST /api/auto_generate/clear 清除预览缓存 ---
@app.post("/api/auto_generate/clear")
def api_auto_generate_clear(req: ClearPreviewRequest):
    """
    清除指定题目的预览缓存（重新出题时调用）。
    不删除磁盘上的题目文件，只清理内存中的未确认预览。
    """
    cleared = False
    if req.question_id in _preview_cache:
        del _preview_cache[req.question_id]
        cleared = True
    return {"success": True, "cleared": cleared}


class SwapSolutionRequest(BaseModel):
    question_id: str


def _build_transfer_preview(cached: Dict) -> Dict:
    """由缓存条目构造换乘题预览（与出题预览同格式，供换方案返回）"""
    segs = cached["segments"]
    seg0, seg1 = segs[0], segs[1]
    stops = cached.get("stops") or []
    first_stop = stops[0]["station_name"] if stops else ""
    last_stop = stops[-1]["station_name"] if stops else ""
    path_desc = (
        f"乘坐 {seg0['train_num']} 从 {seg0['from']} 到 {seg0['to']}（有票），"
        f"换乘 {seg1['train_num']} 从 {seg1['from']} 到 {seg1['to']}（有票）"
    )
    return {
        "question": cached.get("question", ""),
        "question_type": "transfer",
        "target_train_num": cached["target_train_num"],
        "target_train_route": f"{first_stop}→{last_stop}",
        "target_section": f"{seg0['from']}→{seg1['to']}",
        "path_description": path_desc,
        "solution_segments": segs,
    }


def _build_mixed_path_desc(segments: List[Dict]) -> str:
    """混合题各分段叙述（与出题预览共用，供换方案重建预览）"""
    parts = []
    for seg in segments:
        tn = seg["train_num"]
        frm = seg.get("from", "")
        to = seg.get("to", "")
        strat = seg.get("strategy", "direct")
        act_from = seg.get("actual_from", frm)
        act_to = seg.get("actual_to", to)
        if strat == "short_buy":
            parts.append(f"乘坐 {tn} 从 {frm} 到 {to}（有票），补票段 {to}→{act_to} 无余票需上车补票")
        elif strat == "extra_front":
            parts.append(f"乘坐 {tn} 从 {frm} 到 {to} 有票（前额外），可在 {act_from} 站上车，实际乘坐 {act_from}→{act_to}")
        elif strat == "extra_rear":
            parts.append(f"乘坐 {tn} 从 {frm} 到 {to} 有票（后额外），在 {act_to} 站提前下车，实际乘坐 {act_from}→{act_to}")
        else:
            parts.append(f"乘坐 {tn} 从 {frm} 到 {to}（有票）")
    return "，然后 ".join(parts)


def _build_mixed_preview(cached: Dict) -> Dict:
    """由缓存条目构造混合题预览（供换方案返回）"""
    segs = cached["segments"]
    stops = cached.get("stops") or []
    first_stop = stops[0]["station_name"] if stops else ""
    last_stop = stops[-1]["station_name"] if stops else ""
    seg0, segN = segs[0], segs[-1]
    return {
        "question": cached.get("question", ""),
        "question_type": "mixed",
        "target_train_num": cached["target_train_num"],
        "target_train_route": f"{first_stop}→{last_stop}",
        "target_section": f"{seg0['from']}→{segN['to']}",
        "path_description": _build_mixed_path_desc(segs),
        "solution_segments": segs,
    }


def _segment_signature(segments: List[Dict]) -> tuple:
    """判断两套方案是否相同（车次 + 购买区间 + 乘坐区间）"""
    return tuple(
        (s.get("train_num"), s.get("from_station_id"), s.get("to_station_id"),
         s.get("ride_from_station_id"), s.get("ride_to_station_id"))
        for s in segments
    )


def _swap_transfer(cached: Dict, rw_conn: sqlite3.Connection) -> Tuple[List[Dict], Dict]:
    """换乘题换方案：从已枚举的 (M, U) 里重新随机挑一个"""
    solutions = cached.get("solutions") or []
    if len(solutions) < 2:
        raise HTTPException(status_code=400, detail="该车次只有 1 个可行换乘方案，无法再换")
    current = cached.get("chosen_solution") or {}
    remaining = [
        s for s in solutions
        if (s.get("mid_id"), s.get("train_num")) != (current.get("mid_id"), current.get("train_num"))
    ]
    if not remaining:
        raise HTTPException(status_code=400, detail="没有其他可行换乘方案可换")
    s = random.choice(remaining)
    segs = _build_transfer_segments(
        None, rw_conn,
        cached["target_train_num"], cached.get("stops") or [],
        cached["start_station_id"], s["mid_id"], s["train_num"],
        cached["end_station_id"],
        cached.get("seat_type", "class2"), cached.get("people_count", 2),
    )
    return segs, s


def _swap_mixed(cached: Dict, rw_conn: sqlite3.Connection) -> List[Dict]:
    """混合题换方案：同一第一程车 T 重新随机选中间站与各段换乘车次（重跑 mixed 合法解）。

    用内存库重跑（不写真实题库），直到得到与当前不同的方案。
    """
    stops = cached.get("stops") or []
    stop_ids = [s["station_id"] for s in stops]
    if cached["start_station_id"] not in stop_ids:
        raise HTTPException(status_code=400, detail="缓存缺少混合题起点信息，无法换方案")
    from_idx = stop_ids.index(cached["start_station_id"])
    to_idx = len(stops) - 1
    current_sig = _segment_signature(cached.get("segments") or [])
    for _ in range(20):
        q_conn = _create_in_memory_question_db()
        try:
            result = _write_legal_solution(
                q_conn, rw_conn,
                question_type="mixed",
                train_num=cached["target_train_num"],
                stops=stops,
                from_id=cached["start_station_id"], to_id=cached["end_station_id"],
                from_idx=from_idx, to_idx=to_idx,
                segment_plans=cached.get("segment_plans") or [],
                seat_type=cached.get("seat_type", "class2"),
                people_count=cached.get("people_count", 2),
            )
            segs = result["segments"]
        except HTTPException:
            continue  # 该次重跑不可行（_RetryTrainError 等），换一次随机再试
        finally:
            q_conn.close()
        if not segs or _segment_signature(segs) == current_sig:
            continue  # 与当前方案相同，重试
        return segs
    raise HTTPException(status_code=400, detail="未能找到与当前不同的可行混合方案，无法换方案")


# --- POST /api/auto_generate/swap 换方案（不变第一程车 T，换中间站/换乘车次） ---
@app.post("/api/auto_generate/swap")
def api_auto_generate_swap(req: SwapSolutionRequest):
    """
    换方案：保持第一程车 T 不变，重新挑一组合法方案并重建预览。
    - 换乘题（transfer）：从该 T 已枚举的全部可行 (M, U) 里重新随机挑一个
    - 混合题（mixed）：同一 T 重新随机选中间站与各段换乘车次（重跑 mixed 合法解）
    存在性配对（0_/1_ 同 T 同终点）会同步换到同一个新方案，保持两者一致。
    """
    if req.question_id not in _preview_cache:
        raise HTTPException(status_code=400, detail="该题目没有可换方案的预览，请先生成")
    cached = _preview_cache[req.question_id]
    qtype = cached.get("question_type")
    if qtype not in ("transfer", "mixed"):
        raise HTTPException(status_code=400, detail="仅换乘/混合题型支持换方案")

    rw_conn = get_railway_conn()
    try:
        if qtype == "transfer":
            new_segments, chosen = _swap_transfer(cached, rw_conn)
        else:
            new_segments = _swap_mixed(cached, rw_conn)
            chosen = None
    finally:
        rw_conn.close()

    # 更新本 qid 缓存
    cached["segments"] = new_segments
    if chosen is not None:
        cached["chosen_solution"] = chosen
    _preview_cache[req.question_id] = cached

    # 同步存在性配对（同 T 同终点的其他缓存条目，如 0_/1_）→ 保持同一合法解
    preview_builder = _build_transfer_preview if qtype == "transfer" else _build_mixed_preview
    previews = [{"question_id": req.question_id, "preview": preview_builder(cached)}]
    for other_qid, other in list(_preview_cache.items()):
        if other_qid == req.question_id:
            continue
        if (other.get("question_type") == qtype
                and other.get("target_train_num") == cached["target_train_num"]
                and other.get("end_station_id") == cached["end_station_id"]):
            other["segments"] = new_segments
            if chosen is not None:
                other["chosen_solution"] = chosen
            _preview_cache[other_qid] = other
            previews.append({"question_id": other_qid, "preview": preview_builder(other)})

    return {"success": True, "message": "已换方案", "questions": previews}


def _confirm_generated_question(question_id: str,
                                question_type: str = "",
                                answer: str = "",
                                interference: bool = False,
                                interference_density: float = 0.0) -> str:
    """将预览缓存中的题目落盘（DB + metadata），成功后清除缓存。

    手动确认与批量出题共用同一落盘逻辑；失败抛 HTTPException。
    返回确认消息文本。
    """
    if question_id not in _preview_cache:
        raise HTTPException(status_code=400, detail=f"题目 {question_id} 尚未生成预览，请先生成")

    cached = _preview_cache[question_id]

    # 检查磁盘是否已存在同名 DB（防止并发冲突）
    db_path = get_question_db_path(question_id)
    if os.path.exists(db_path):
        del _preview_cache[question_id]
        raise HTTPException(status_code=400, detail=f"题目 {question_id} 的数据库文件已存在")

    # 创建真实数据库文件
    create_question_db(question_id)
    q_conn = sqlite3.connect(db_path)
    try:
        for seg in cached["segments"]:
            db_update_ticket(
                q_conn, seg["train_num"],
                seg["from_station_id"],
                seg["to_station_id"],
                seg.get("seat_type", "class2"),
                seg["tickets"]
            )

        # 干扰票处理：与预览（_generate_one_variant）保持同一套全局池逻辑
        if cached.get("random_tickets"):
            solution_pairs = set()
            for seg in cached["segments"]:
                solution_pairs.add((seg["from_station_id"], seg["to_station_id"]))
            fake_mode = bool(cached.get("fake_interference"))
            from_id = cached.get("start_station_id")
            to_id = cached.get("end_station_id")
            if fake_mode:
                block_pairs = set()
                guard_from = guard_to = None
            else:
                block_pairs = {(from_id, to_id)}
                guard_from, guard_to = (
                    (from_id, to_id) if cached.get("question_type") in ("transfer", "mixed")
                    else (None, None)
                )
            # 需要 rw_conn 读取全部车次信息（内部一次性缓存经停站）
            rw_conn = get_railway_conn()
            try:
                _add_interference_all_trains(
                    q_conn, rw_conn, cached["target_train_num"],
                    solution_pairs, cached["interference_density"],
                    block_pairs=block_pairs, fake=fake_mode,
                    people_count=cached.get("people_count") or 2,
                    shortbuy_guard_from=guard_from, shortbuy_guard_to=guard_to,
                )
            finally:
                rw_conn.close()
            if not fake_mode:
                _clear_direct_route(q_conn, cached["target_train_num"], from_id, to_id)

        q_conn.commit()
    except Exception as e:
        q_conn.close()
        # 写入失败则清理
        if os.path.exists(db_path):
            os.remove(db_path)
        del _preview_cache[question_id]
        raise HTTPException(status_code=500, detail=f"写入数据库失败: {str(e)}")
    q_conn.close()

    # 收集 trains 列表
    trains_list = sorted({seg["train_num"] for seg in cached["segments"]})

    meta_kwargs = {
        "question_id": question_id,
        "status": "completed",
        "train_count": len(trains_list),
        "trains": trains_list,
        "source": "auto",
        "question_type": question_type or cached["question_type"],
        "answer": answer or None,
        "question": cached.get("question"),
        "segment_plans": cached.get("segment_plans"),
        "interference": cached.get("interference"),
        "interference_mode": (
            "fake" if cached.get("fake_interference")
            else ("real" if cached.get("interference") else None)
        ),
        "question_mode": cached.get("question_mode"),
        "people_count": cached.get("people_count"),
        "seat_type": cached.get("seat_type"),
        "start_station_id": cached.get("start_station_id"),
        "end_station_id": cached.get("end_station_id"),
        "criterion": cached.get("criterion", "comprehensive"),  # 评判标准（仅选择性）
        "ground_truth": cached.get("segments"),   # 结构化标答存进 metadata（含购买+乘坐区间 id）
    }
    # 仅选择性题（有干扰票）才记录干扰密度，存在性题不写入该字段
    if cached.get("interference"):
        meta_kwargs["interference_density"] = cached.get("interference_density")
    # 行为约束：仅选择性题且非空时落盘（纯要求，无标准答案）
    if cached.get("constraints"):
        meta_kwargs["constraints"] = cached.get("constraints")
    update_question_metadata(**meta_kwargs)

    # 清除缓存
    del _preview_cache[question_id]

    return f"题目 {question_id} 已确认生成，可在测试器中加载使用"


# --- POST /api/auto_generate/confirm 确认生成 ---
@app.post("/api/auto_generate/confirm")
def api_auto_generate_confirm(req: ConfirmAutoGenerateRequest):
    """确认生成自动出题的题目（落盘 DB + metadata）。"""
    message = _confirm_generated_question(
        req.question_id, req.question_type, req.answer,
        req.interference, req.interference_density,
    )
    return {"success": True, "message": message}


# --- GET /api/question/list 列出所有题目 ---
def _ground_truth_summary(gt) -> str:
    """结构化标答（ground_truth）→ 文本摘要，供题目管理「标准答案」列预览。

    示例：G2689 淮安东→郑州东（乘 HAU→ZAF） class0×5；G3185 ...
    购买区间与乘坐区间一致时不显示乘坐括号。
    """
    if not gt:
        return ""
    parts = []
    for seg in gt:
        if not isinstance(seg, dict):
            continue
        ride = ""
        rf = seg.get("ride_from_station_id") or seg.get("ride_from") or ""
        rt = seg.get("ride_to_station_id") or seg.get("ride_to") or ""
        if rf and rt and (rf != seg.get("from_station_id") or rt != seg.get("to_station_id")):
            ride = f"（乘 {rf}→{rt}）"
        seat = seg.get("seat_type") or ""
        tickets = seg.get("tickets")
        parts.append(
            f"{seg.get('train_num', '')} {seg.get('from', '')}→{seg.get('to', '')}{ride}"
            + (f" {seat}×{tickets}" if tickets is not None else "")
        )
    return "；".join(parts)


@app.get("/api/question/list")
def api_question_list(
    status_filter: Optional[str] = Query(None),
    source_filter: Optional[str] = Query(None, alias="source"),
    type_filter: Optional[str] = Query(None, alias="type"),
    keyword: Optional[str] = Query(None),
):
    """列出所有题目，支持筛选"""
    metadata = load_metadata()

    from config import QUESTION_DIR
    db_files = {}
    if os.path.exists(QUESTION_DIR):
        for f in os.listdir(QUESTION_DIR):
            if f.endswith(".db"):
                qid = f[:-3]
                db_files[qid] = True

    result = []

    # 收集有 metadata 的题目
    seen = set()
    for qid, meta in metadata.items():
        seen.add(qid)
        if status_filter and meta.get("status") != status_filter:
            continue
        if source_filter and meta.get("source") != source_filter:
            continue
        if type_filter and meta.get("type") != type_filter:
            continue
        if keyword and keyword.lower() not in qid.lower():
            continue
        result.append({
            "question_id": qid,
            "status": meta.get("status", "unknown"),
            "source": meta.get("source", ""),
            "type": meta.get("type", ""),
            "question_type": meta.get("question_type", ""),
            "train_count": meta.get("train_count", 0),
            # 标准答案：优先旧 answer 文本字段；自动出题无该字段时用结构化标答 ground_truth 生成摘要
            "answer": meta.get("answer") or _ground_truth_summary(meta.get("ground_truth")),
            "question": meta.get("question", ""),
            "nl_question": meta.get("nl_question", ""),
            "db_exists": qid in db_files,
        })

    # 也列出孤立 .db 文件（有 db 无 metadata），供清理
    for qid in db_files:
        if qid not in seen:
            if keyword and keyword.lower() not in qid.lower():
                continue
            result.append({
                "question_id": qid,
                "status": "orphan",
                "source": "",
                "type": "",
                "train_count": 0,
                "answer": "",
                "db_exists": True,
            })

    def _natural_key(qid: str):
        """统一题号自然序：纯数字(空前缀)最前 → 0_/1_/2_（前缀+尾随数字自然序）→ 字母前缀 → 其他。
        例：123 → 0_2 → 0_10 → 1_2 → 1_10 → 1_26 → a1 …"""
        s = str(qid)
        if re.fullmatch(r"\d+", s):          # 纯数字（空前缀）最前
            return (0, "", int(s))
        m = re.match(r"^(.*?)(\d+)$", s)     # 前缀 + 尾随数字：0_2 / 1_10 / a1
        if m:
            prefix = m.group(1)
            # 前缀含数字（0_/1_/2_）排在字母前缀之前
            group = 1 if re.search(r"\d", prefix) else 2
            return (group, prefix, int(m.group(2)))
        return (3, s, 0)                      # 其他（无尾随数字）最后
    result.sort(key=lambda x: _natural_key(x["question_id"]))
    return {"questions": result, "total": len(result)}


class InitQuestionRequest(BaseModel):
    question_id: str
    train_num: str


# --- POST /api/question/init 初始化题目余票数据库 ---
@app.post("/api/question/init")
def api_question_init(req: InitQuestionRequest):
    """
    初始化题目中指定车次的余票数据库。
    - 如果题目数据库不存在，自动创建
    - 每次调用都将车次加入 metadata 的 trains 列表（不重置已有数据）
    """
    db_path = get_question_db_path(req.question_id)
    if not os.path.exists(db_path):
        create_question_db(req.question_id)

    rw_conn = get_railway_conn()
    q_conn = sqlite3.connect(db_path)
    try:
        # 校验请求的车次存在
        if not validate_train_exists(rw_conn, req.train_num):
            raise HTTPException(status_code=400, detail=f"车次 {req.train_num} 不存在")

        # 收集要加载的所有车次（主车次 + 同一物理列车的关联车次）
        all_trains = [req.train_num]
        same_map = load_same_train_map()
        associated = same_map.get(req.train_num, [])
        # 只添加存在的关联车次
        for tn in associated:
            if validate_train_exists(rw_conn, tn) and tn not in all_trains:
                all_trains.append(tn)

        # 记录已加载车次到元数据（合并到已有列表，不重置）
        update_question_metadata(
            req.question_id, status="draft",
            trains=all_trains
        )
        msg = f"车次 {req.train_num} 已加入题目 {req.question_id}"
        if associated:
            msg += f"，同时添加了同车关联号 {', '.join(associated)}"
        return {"success": True, "message": msg}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        rw_conn.close()
        q_conn.close()


# --- GET /api/question/{question_id}/exists 检查题目是否存在 ---
@app.get("/api/question/{question_id}/exists")
def api_question_exists(question_id: str):
    """检查题目数据库文件是否已存在"""
    db_path = get_question_db_path(question_id)
    exists = os.path.exists(db_path)
    return {"question_id": question_id, "exists": exists}


# --- GET /api/question/{question_id}/trains 获取题目中已初始化的车次列表 ---
@app.get("/api/question/{question_id}/trains")
def api_question_trains(question_id: str):
    """返回指定题目已加载的车次列表（从元数据读取）"""
    meta = get_question_metadata(question_id)
    trains_list = meta.get("trains", [])
    trains = [{"train_num": tn} for tn in trains_list]
    return {"question_id": question_id, "trains": trains, "count": len(trains)}


# ============================================================
# 第三部分：测试器接口
# ============================================================

# 对话状态（内存中维护）
# 每个用户会话的对话历史和题目状态
chat_sessions: Dict[str, List[Dict]] = {}
# 会话元数据：cumulative token_usage 和 duration
chat_session_meta: Dict[str, Dict] = {}
# 会话使用的模型名称
chat_session_model: Dict[str, str] = {}
# 会话工具调用日志（供测试完成落盘，统计 avg_tool_calls 用）
chat_session_tool_calls: Dict[str, List[Dict]] = {}
api_keys: Dict[str, str] = {}


class CompleteQuestionRequest(BaseModel):
    question_id: str


# --- DELETE /api/question/{question_id}/train/{train_num} 删除车次 ---
@app.delete("/api/question/{question_id}/train/{train_num}")
def api_question_delete_train(question_id: str, train_num: str):
    """从题目中删除指定车次（metadata + 数据库余票数据）"""
    db_path = get_question_db_path(question_id)
    if not os.path.exists(db_path):
        raise HTTPException(status_code=404, detail=f"题目 {question_id} 不存在")

    q_conn = sqlite3.connect(db_path)
    try:
        # 删除数据库中的余票数据
        delete_train_tickets(q_conn, train_num)
        # 从 metadata 中移除车次
        remove_train_from_metadata(question_id, train_num)
        return {"success": True, "message": f"车次 {train_num} 已从题目 {question_id} 中删除"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        q_conn.close()


# --- POST /api/question/complete 标记题目完成 ---
@app.post("/api/question/complete")
def api_question_complete(req: CompleteQuestionRequest):
    """标记题目为 completed 状态，供测试器加载"""
    db_path = get_question_db_path(req.question_id)
    if not os.path.exists(db_path):
        raise HTTPException(status_code=404, detail=f"题目 {req.question_id} 不存在，请先加载车次")

    # 从 metadata 读取实际车次数，确保 train_count 正确
    meta = get_question_metadata(req.question_id)
    actual_train_count = len(meta.get("trains", []))
    update_question_metadata(
        req.question_id, status="completed",
        train_count=actual_train_count
    )
    return {"success": True, "message": f"题目 {req.question_id} 已完成"}


# --- DELETE /api/question/{question_id} 删除题目 ---
@app.delete("/api/question/{question_id}")
def api_question_delete(question_id: str):
    """删除指定题目（metadata + .db文件 + 缓存中的预览）"""
    # 先清除预览缓存（如果有）
    _preview_cache.pop(question_id, None)

    # 先删除 .db 文件（防止文件被占用导致 metadata 已删但 db 残留）
    db_path = get_question_db_path(question_id)
    db_deleted = False
    if os.path.exists(db_path):
        try:
            os.remove(db_path)
            db_deleted = True
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"删除数据库文件失败: {str(e)}")

    # 再更新 metadata
    metadata = load_metadata()
    if question_id in metadata:
        del metadata[question_id]
        save_metadata(metadata)
        return {"success": True, "message": f"题目 {question_id} 已删除"}

    if db_deleted:
        return {"success": True, "message": f"已清理孤立数据库文件 {question_id}"}

    # 既无 metadata 也无 .db 文件
    raise HTTPException(status_code=404, detail=f"题目 {question_id} 不存在")


class ChatRequest(BaseModel):
    message: str
    # 默认一律空串：空值才会触发 _resolve_llm_config 回落 .env（TEST_MODEL/TEST_API_KEY/DEFAULT_BASE_URL）。
    # 严禁写平台默认值（如 https://api.deepseek.com）——非空默认会挡住回落，导致用 A 平台 key 调 B 平台报 401。
    model_name: str = ""
    api_key: str = ""
    api_base_url: str = ""
    question_id: str = ""
    session_id: str = "default"
    max_iterations: int = 100


class TestCompleteRequest(BaseModel):
    session_id: str = "default"


def _resolve_llm_config(req) -> tuple:
    """ChatRequest 未填 model/key/base_url 时，回落到 .env 的默认配置。

    优先级：请求显式值 > .env（TEST_MODEL → DEFAULT_MODEL / TEST_API_KEY / DEFAULT_BASE_URL）。
    模型名与接口地址不内置任何平台默认（不一定调用 deepseek）：未配置时直接 400 报错。
    """
    from config import ENV, ensure_env_fresh
    ensure_env_fresh()  # .env 被修改后自动重载，避免进程继续用旧 key/旧模型名
    model = (req.model_name or "").strip() or ENV.get("TEST_MODEL", "") or ENV.get("DEFAULT_MODEL", "") or ""
    key = (req.api_key or "").strip() or ENV.get("TEST_API_KEY", "") or ""
    base_url = (req.api_base_url or "").strip() or ENV.get("TEST_BASE_URL", "") or ENV.get("DEFAULT_BASE_URL", "") or ""
    missing = [name for name, val in
               [("模型名（TEST_MODEL / DEFAULT_MODEL）", model),
                ("API Key（TEST_API_KEY）", key),
                ("接口地址（TEST_BASE_URL / DEFAULT_BASE_URL）", base_url)] if not val]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f".env 模型配置缺失：{'、'.join(missing)}。请在根目录 .env 中填写（系统不默认指向任何平台）"
        )
    return model, key, base_url


# --- GET /api/env/model 返回 .env 中解析后的测试模型名（供前端顶栏/输入框显示，不含 key） ---
@app.get("/api/env/model")
def api_env_model():
    from config import ENV, ensure_env_fresh
    ensure_env_fresh()  # .env 被修改后自动重载
    test_model = (ENV.get("TEST_MODEL") or ENV.get("DEFAULT_MODEL") or "").strip()
    return {"test_model": test_model}


@app.post("/api/test/chat")
async def api_test_chat(req: ChatRequest):
    """发送对话消息，调用大模型并处理工具调用循环"""
    model_name, api_key, api_base_url = _resolve_llm_config(req)

    # 设置当前题目
    if req.question_id:
        set_current_question(req.question_id)

    # 初始化或获取会话历史
    if req.session_id not in chat_sessions:
        chat_sessions[req.session_id] = []

    # 每次发消息都更新模型名称（历史记录中可能被 resetChat 删除）
    chat_session_model[req.session_id] = model_name

    messages = chat_sessions[req.session_id]

    # 如果这是第一条消息，添加系统提示词（最大工具调用次数由下方循环 max_iterations 限制）
    if not messages:
        messages.append({"role": "system", "content": SYSTEM_PROMPT})

    # 添加用户消息
    messages.append({"role": "user", "content": req.message})

    # 工具调用循环
    tool_calls_log = []
    token_usage = {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0}
    last_prompt_tokens = 0  # 上轮 prompt_tokens，用于计算增量
    start_time = datetime.now()

    try:
        import httpx

        # 构造请求体
        request_body = {
            "model": model_name,
            "messages": messages,
            "tools": TOOLS,
            "tool_choice": "auto",
            "stream": False,
            "temperature": TEST_TEMPERATURE,  # 测试器对话温度（config.py）
        }

        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}",
        }

        max_iterations = req.max_iterations  # 防止无限循环
        iteration = 0
        final_content = ""
        final_reasoning = ""
        # 批量测试会话（batch_test_<qid>）：逐轮写入批量日志，供前端实时观察执行进度
        batch_qid = req.session_id[len("batch_test_"):] if req.session_id.startswith("batch_test_") else None

        async with httpx.AsyncClient(timeout=60) as client:
            while iteration < max_iterations:
                iteration += 1
                if batch_qid:
                    _batch_log("batch_test", f"{batch_qid} 第 {iteration}/{max_iterations} 轮对话：请求模型 {model_name}")

                # 调用大模型 API
                api_url = f"{api_base_url.rstrip('/')}/chat/completions"
                response = await client.post(api_url, json=request_body, headers=headers)
                response_data = response.json()

                if "error" in response_data:
                    return {"error": response_data["error"].get("message", "API 调用失败")}

                choice = response_data["choices"][0]
                message = choice["message"]

                # 记录 token 用量（仅计新增 prompt tokens，扣除缓存命中）
                if "usage" in response_data:
                    usage = response_data["usage"]
                    current_prompt = usage.get("prompt_tokens", 0)
                    # 减去缓存命中的 tokens
                    prompt_details = usage.get("prompt_tokens_details")
                    if isinstance(prompt_details, dict):
                        cached = prompt_details.get("cached_tokens", 0)
                        current_prompt = max(0, current_prompt - cached)
                    delta = max(0, current_prompt - last_prompt_tokens)
                    token_usage["prompt_tokens"] += delta
                    last_prompt_tokens = current_prompt
                    token_usage["completion_tokens"] += usage.get("completion_tokens", 0)
                    token_usage["total_tokens"] = token_usage["prompt_tokens"] + token_usage["completion_tokens"]

                # 提取 reasoning_content（思考链）
                reasoning = message.get("reasoning_content") or choice.get("delta", {}).get("reasoning_content") or ""

                # 检查是否有工具调用
                tool_calls = message.get("tool_calls")
                if not tool_calls:
                    # 模型给出最终回答
                    final_content = message.get("content", "")
                    messages.append({
                        "role": "assistant",
                        "content": final_content,
                    })
                    break

                # 处理工具调用
                assistant_msg = {"role": "assistant", "content": message.get("content", "")}
                assistant_msg["tool_calls"] = []
                messages.append(assistant_msg)

                for tc in tool_calls:
                    func_name = tc["function"]["name"]
                    func_args_str = tc["function"]["arguments"]
                    try:
                        func_args = json.loads(func_args_str) if isinstance(func_args_str, str) else func_args_str
                    except json.JSONDecodeError:
                        func_args = {}

                    # 执行工具（传入题目 ID：并发批量测试时避免全局状态串题）
                    tool_result = await execute_tool_handler(func_name, func_args, req.question_id)
                    if batch_qid:
                        _batch_log("batch_test", f"{batch_qid} 第 {iteration} 轮执行工具 {func_name}")

                    tool_calls_log.append({
                        "tool_name": func_name,
                        "arguments": func_args,
                        "result": tool_result,
                    })

                    messages.append({
                        "role": "tool",
                        "tool_call_id": tc["id"],
                        "content": json.dumps(tool_result, ensure_ascii=False),
                    })

                    assistant_msg["tool_calls"].append({
                        "id": tc["id"],
                        "type": "function",
                        "function": {"name": func_name, "arguments": func_args_str},
                    })

                # 更新请求体中的消息
                request_body["messages"] = messages

        duration = (datetime.now() - start_time).total_seconds()

        # 累计保存到会话元数据
        sid = req.session_id
        if sid not in chat_session_meta:
            chat_session_meta[sid] = {"token_usage": {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0}, "duration": 0}
        meta = chat_session_meta[sid]
        if req.question_id:
            meta["question_id"] = req.question_id
        meta["token_usage"]["prompt_tokens"] += token_usage.get("prompt_tokens", 0)
        meta["token_usage"]["completion_tokens"] += token_usage.get("completion_tokens", 0)
        meta["token_usage"]["total_tokens"] += token_usage.get("total_tokens", 0)
        meta["duration"] += duration

        # 累计工具调用日志（供 api_test_complete 落盘）
        chat_session_tool_calls[sid] = chat_session_tool_calls.get(sid, []) + tool_calls_log

        return {
            "reply": final_content,
            "reasoning": final_reasoning,
            "tool_calls": tool_calls_log,
            "token_usage": token_usage,
            "duration": duration,
            "iteration_count": iteration,
        }

    except Exception as e:
        logger.error(f"测试对话出错: {e}")
        return {"error": str(e)}


@app.post("/api/test/chat/stream")
async def api_test_chat_stream(req: ChatRequest, request: Request):
    """SSE 流式对话"""
    model_name, api_key, api_base_url = _resolve_llm_config(req)

    # 设置当前题目
    if req.question_id:
        set_current_question(req.question_id)

    # 初始化或获取会话历史
    if req.session_id not in chat_sessions:
        chat_sessions[req.session_id] = []

    # 每次发消息都更新模型名称（历史记录中可能被 resetChat 删除）
    chat_session_model[req.session_id] = model_name

    messages = chat_sessions[req.session_id]

    # 如果这是第一条消息，添加系统提示词（最大工具调用次数由下方循环 max_iterations 限制）
    if not messages:
        messages.append({"role": "system", "content": SYSTEM_PROMPT})

    # 添加用户消息
    messages.append({"role": "user", "content": req.message})

    # 累计 token 用量（仅计新增，每轮减去上轮已计的历史）
    token_usage = {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0}
    last_prompt_tokens = 0  # 上轮返回的 prompt_tokens，用于计算增量
    start_time = datetime.now()

    async def event_generator():
        nonlocal token_usage, last_prompt_tokens
        import httpx

        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}",
        }

        api_url = f"{api_base_url.rstrip('/')}/chat/completions"

        # 全量累加器（跨迭代持久化，用于 done 事件输出完整文本）
        full_content_accumulated = ""
        full_reasoning_accumulated = ""

        try:
            async with httpx.AsyncClient(timeout=120) as client:
                request_body = {
                    "model": model_name,
                    "messages": messages,
                    "tools": TOOLS,
                    "tool_choice": "auto",
                    "stream": True,
                    "temperature": TEST_TEMPERATURE,  # 测试器对话温度（config.py）
                }

                max_iterations = req.max_iterations
                iteration = 0

                while iteration < max_iterations:
                    iteration += 1

                    # —— 安全检测：客户端断连 ——
                    if await request.is_disconnected():
                        logger.info(f"客户端断开连接，停止循环 (session={req.session_id})")
                        return

                    accumulated_content = ""
                    accumulated_reasoning = ""
                    tool_calls_buffer = {}

                    try:
                        # 流式调用 LLM
                        async with client.stream("POST", api_url, json=request_body, headers=headers) as resp:
                            if resp.status_code != 200:
                                error_text = await resp.aread()
                                yield f"data: {json.dumps({'type': 'error', 'content': f'API 调用失败 ({resp.status_code}): {str(error_text)[:200]}'})}\n\n"
                                return

                            async for line in resp.aiter_lines():
                                if not line.startswith("data: "):
                                    continue
                                data_str = line[6:]
                                if data_str.strip() == "[DONE]":
                                    break
                                try:
                                    chunk = json.loads(data_str)
                                    # 捕获 token 用量（仅累加每轮新增的 prompt tokens，扣除缓存命中）
                                    if "usage" in chunk and chunk["usage"] is not None:
                                        current_prompt = chunk["usage"].get("prompt_tokens", 0)
                                        # 减去缓存命中的 tokens
                                        prompt_details = chunk["usage"].get("prompt_tokens_details")
                                        if isinstance(prompt_details, dict):
                                            cached = prompt_details.get("cached_tokens", 0)
                                            current_prompt = max(0, current_prompt - cached)
                                        delta = max(0, current_prompt - last_prompt_tokens)
                                        token_usage["prompt_tokens"] += delta
                                        last_prompt_tokens = current_prompt
                                        token_usage["completion_tokens"] += chunk["usage"].get("completion_tokens", 0)
                                        token_usage["total_tokens"] = token_usage["prompt_tokens"] + token_usage["completion_tokens"]
                                    if "choices" not in chunk or not chunk["choices"]:
                                        continue
                                    delta = chunk["choices"][0].get("delta", {})

                                    # 思考过程（同时累加到轮次变量和全量变量）
                                    if delta.get("reasoning_content"):
                                        rc = delta["reasoning_content"]
                                        accumulated_reasoning += rc
                                        full_reasoning_accumulated += rc
                                        yield f"data: {json.dumps({'type': 'reasoning', 'content': rc})}\n\n"

                                    # 内容 token（同时累加到轮次变量和全量变量）
                                    if delta.get("content"):
                                        c = delta["content"]
                                        accumulated_content += c
                                        full_content_accumulated += c
                                        yield f"data: {json.dumps({'type': 'token', 'content': c})}\n\n"

                                    # 工具调用
                                    if delta.get("tool_calls"):
                                        for tc in delta["tool_calls"]:
                                            idx = tc["index"]
                                            if idx not in tool_calls_buffer:
                                                tool_calls_buffer[idx] = {"id": "", "function": {"name": "", "arguments": ""}}
                                            if tc.get("id"):
                                                tool_calls_buffer[idx]["id"] = tc["id"]
                                            if tc.get("function"):
                                                fn = tc["function"]
                                                if fn.get("name"):
                                                    tool_calls_buffer[idx]["function"]["name"] = fn["name"]
                                                if fn.get("arguments"):
                                                    tool_calls_buffer[idx]["function"]["arguments"] += fn["arguments"]

                                except json.JSONDecodeError:
                                    pass
                    except Exception as stream_err:
                        logger.error(f"流式请求出错: {stream_err}")
                        yield f"data: {json.dumps({'type': 'error', 'content': str(stream_err)})}\n\n"
                        return

                    # 如果本轮有工具调用
                    if tool_calls_buffer:
                        tool_calls_list = [v for k, v in sorted(tool_calls_buffer.items())]

                        # 发送工具调用事件（前台渲染折叠卡片）
                        yield f"data: {json.dumps({'type': 'tool_call', 'tool_calls': [(tc['function']['name'], tc['function'].get('arguments', '{}')) for tc in tool_calls_list]})}\n\n"

                        # 保存 assistant 消息（reasoning 不写入 messages，避免回流给模型）
                        assistant_msg = {
                            "role": "assistant",
                            "content": accumulated_content,
                            "tool_calls": [],
                        }
                        messages.append(assistant_msg)

                        for tc in tool_calls_list:
                            try:
                                func_args = json.loads(tc["function"]["arguments"])
                            except (json.JSONDecodeError, TypeError):
                                func_args = {}

                            # 执行工具（传入题目 ID：并发批量测试时避免全局状态串题）
                            tool_result = await execute_tool_handler(tc["function"]["name"], func_args, req.question_id)

                            # 累计工具调用日志（供 api_test_complete 落盘）
                            chat_session_tool_calls.setdefault(req.session_id, []).append({
                                "tool_name": tc["function"]["name"],
                                "arguments": func_args,
                                "result": tool_result,
                            })

                            # 发送 tool_result 事件，让前端展示工具返回结果
                            yield f"data: {json.dumps({
                                'type': 'tool_result',
                                'tool_name': tc['function']['name'],
                                'arguments': tc['function'].get('arguments', '{}'),
                                'result': tool_result,
                            })}\n\n"

                            messages.append({
                                "role": "tool",
                                "tool_call_id": tc["id"],
                                "content": json.dumps(tool_result, ensure_ascii=False),
                            })

                            assistant_msg["tool_calls"].append({
                                "id": tc["id"],
                                "type": "function",
                                "function": {"name": tc["function"]["name"], "arguments": tc["function"]["arguments"]},
                            })

                        # 更新请求体，继续下一轮
                        request_body["messages"] = messages
                        accumulated_content = ""
                        accumulated_reasoning = ""
                        tool_calls_buffer = {}
                        continue  # 下一轮迭代

                    else:
                        # 没有工具调用 → 最终回复
                        duration = (datetime.now() - start_time).total_seconds()

                        # 若 API 未返回 usage（部分模型流式不支持），按字符数估算 token
                        if token_usage["total_tokens"] == 0:
                            # 估算：中文字符≈2 tokens，英文≈0.3 tokens
                            total_chars = len(full_content_accumulated) + len(full_reasoning_accumulated)
                            total_chars += sum(len(m.get("content","")) for m in messages[-3:] if m.get("role")=="user")
                            estimated = int(total_chars * 0.5) + 10  # 加 buffer
                            token_usage["prompt_tokens"] = max(token_usage["prompt_tokens"], estimated // 2)
                            token_usage["completion_tokens"] = max(token_usage["completion_tokens"], estimated // 2)
                            token_usage["total_tokens"] = max(token_usage["total_tokens"], estimated)

                        # 更新会话元数据
                        sid = req.session_id
                        if sid not in chat_session_meta:
                            chat_session_meta[sid] = {
                                "token_usage": {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0},
                                "duration": 0,
                            }
                        meta = chat_session_meta[sid]
                        meta["token_usage"]["prompt_tokens"] += token_usage.get("prompt_tokens", 0)
                        meta["token_usage"]["completion_tokens"] += token_usage.get("completion_tokens", 0)
                        meta["token_usage"]["total_tokens"] += token_usage.get("total_tokens", 0)
                        meta["duration"] += duration

                        # 保存到对话历史（reasoning 不写入 messages）
                        messages.append({
                            "role": "assistant",
                            "content": accumulated_content,
                        })

                        # 发送完成事件（使用全量累加器，包含所有轮次内容）
                        yield f"data: {json.dumps({
                            'type': 'done',
                            'content': full_content_accumulated,
                            'token_usage': token_usage,
                            'duration': duration,
                        })}\n\n"
                        return

                # 超限
                yield f"data: {json.dumps({'type': 'error', 'content': '工具调用循环超过最大次数'})}\n\n"

        except Exception as e:
            logger.error(f"测试流式对话出错: {e}")
            yield f"data: {json.dumps({'type': 'error', 'content': str(e)})}\n\n"

    return StreamingResponse(event_generator(), media_type="text/event-stream")


def _parse_ai_final_plan(assistant_content: str) -> Optional[List[Dict]]:
    """
    从模型回复末尾解析 JSON 格式的 final_plan。
    格式: {"final_plan": [{"train_num":"G1","from":"VNP","to":"JGK","seat_type":"class2","tickets":2}]}
    尝试匹配 ```json ... ```、``` ... ``` 或裸 JSON（正确处理嵌套花括号）。
    """
    if not assistant_content:
        return None
    import re
    json_str = None

    # 1. 尝试 ```json ... ``` 或 ``` ... ```
    match = re.search(r'```(?:json)?\s*(\{.*?\})\s*```', assistant_content, re.DOTALL)
    if match:
        json_str = match.group(1)
    else:
        # 2. 定位 "final_plan" 所在的最外层花括号
        idx = assistant_content.find('"final_plan"')
        if idx >= 0:
            # 向左找第一个 {
            start = assistant_content.rfind('{', 0, idx)
            if start >= 0:
                # 向右逐字符匹配花括号，找到对应闭合 }
                depth = 0
                for i in range(start, len(assistant_content)):
                    if assistant_content[i] == '{':
                        depth += 1
                    elif assistant_content[i] == '}':
                        depth -= 1
                        if depth == 0:
                            json_str = assistant_content[start:i + 1]
                            break
    if not json_str:
        return None  # 没有找到 final_plan JSON
    try:
        data = json.loads(json_str)
        plan = data.get("final_plan", [])
        if not isinstance(plan, list):
            return None
        if len(plan) == 0:
            return []  # 模型明确输出无解
        # 用 verifier 的统一字段契约归一化（兼容 from/to 与 from_station_id/to_station_id）
        validated = [e for e in normalize_final_plan(plan) if not e.get("__invalid__")]
        if validated:
            return validated
        # 有 final_plan JSON 但无有效条目
        return []
    except (json.JSONDecodeError, ValueError, TypeError):
        pass
    return None


@app.post("/api/test/complete")
def api_test_complete(req: TestCompleteRequest):
    """标记测试完成，保存记录"""
    session = chat_sessions.get(req.session_id)
    if not session:
        raise HTTPException(status_code=400, detail="会话不存在")

    # 提取对话信息
    user_input = ""
    final_answer = ""
    for msg in reversed(session):
        if msg["role"] == "user":
            user_input = msg["content"]
            break

    for msg in reversed(session):
        if msg["role"] == "assistant" and msg.get("content"):
            final_answer = msg["content"]
            break

    # 获取会话累计的 token_usage 和 duration
    meta = chat_session_meta.get(req.session_id, {})
    token_usage = meta.get("token_usage", {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0})
    duration = meta.get("duration", 0)

    # 题目 ID：优先取会话记录的（并发批量测试不依赖全局状态），兜底全局当前题目
    question_id = meta.get("question_id")
    if not question_id:
        try:
            question_id = get_current_question()
        except ValueError:
            question_id = "unknown"

    # 获取模型名称
    model_name = chat_session_model.get(req.session_id, "unknown")

    # 获取会话工具调用日志
    tool_calls = chat_session_tool_calls.get(req.session_id, [])

    # 生成保存路径
    from config import LOGS_TEST_DIR
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{timestamp}_{model_name}_{question_id}.json"
    filepath = os.path.join(LOGS_TEST_DIR, filename)

    # 从模型回复中解析 AI 输出的结构化 final_plan
    final_plan = _parse_ai_final_plan(final_answer)
    if final_plan is None:
        final_plan = []  # 无 JSON 输出视为未规划
        plan_status = "no_plan"  # 模型未输出 final_plan
    elif len(final_plan) == 0:
        plan_status = "empty_plan"  # 模型明确无解（与 verifier verdict 词汇统一）
    else:
        plan_status = "has_solution"  # 有方案

    record = {
        "timestamp": datetime.now().isoformat(),
        "question_id": question_id,
        "model_name": model_name,
        "user_input": user_input,
        "conversation": session,
        "final_answer": final_answer,
        "final_plan": final_plan,
        "plan_status": plan_status,
        "token_usage": token_usage,
        "duration": duration,
        "tool_calls": tool_calls,
    }

    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(record, f, ensure_ascii=False, indent=2)

    return {
        "success": True,
        "filepath": filepath,
        "filename": filename,
        "message": f"测试记录已保存至 {filename}",
        "token_usage": token_usage,
        "duration": duration,
        "model_name": model_name,
        "final_plan": final_plan,
        "plan_status": plan_status,
        "tool_calls": tool_calls,
    }


class TestResetRequest(BaseModel):
    session_id: str = "default"


@app.post("/api/test/reset")
def api_test_reset(req: TestResetRequest):
    """重置对话"""
    session_id = req.session_id
    chat_sessions[session_id] = []
    if session_id in chat_session_meta:
        del chat_session_meta[session_id]
    if session_id in chat_session_model:
        del chat_session_model[session_id]
    if session_id in chat_session_tool_calls:
        del chat_session_tool_calls[session_id]
    return {"success": True, "message": "对话已重置"}


async def execute_tool_handler(func_name: str, func_args: Dict, question_id: Optional[str] = None) -> Any:
    """执行工具函数

    question_id：当前题目 ID（余票查询用）。并发批量测试时按调用传入以避免全局状态串题；
    传 None 时回落全局 get_current_question()（单会话手动测试兼容）。"""
    rw_conn = get_railway_conn()
    try:
        if func_name == "query_trains_between_stations":
            data = get_routes_between(
                rw_conn,
                func_args.get("from_station_id", ""),
                func_args.get("to_station_id", "")
            )
            return data

        elif func_name == "query_trains_at_station":
            result = get_station_trains(rw_conn, func_args.get("station_id", ""))
            if result is None:
                return {"error": f"车站 {func_args.get('station_id', '')} 未找到"}
            return result["trains"]

        elif func_name == "query_train_detail":
            stops = get_train_stops(rw_conn, func_args.get("train_num", ""))
            if not stops:
                return {"error": f"车次 {func_args.get('train_num', '')} 未找到或无经停数据"}
            return {"stops": stops}

        elif func_name == "list_stations":
            keyword = func_args.get("keyword", "")
            limit = max(1, min(int(func_args.get("limit", 50) or 50), 200))
            page = max(1, int(func_args.get("page", 1) or 1))
            cursor = rw_conn.cursor()
            offset = (page - 1) * limit
            if keyword:
                cursor.execute(
                    "SELECT station_id, station_name FROM stations WHERE station_name LIKE ? OR station_id LIKE ? ORDER BY station_name LIMIT ? OFFSET ?",
                    (f"%{keyword}%", f"%{keyword}%", limit, offset)
                )
            else:
                cursor.execute(
                    "SELECT station_id, station_name FROM stations ORDER BY station_name LIMIT ? OFFSET ?",
                    (limit, offset)
                )
            data = [{"station_id": row[0], "station_name": row[1]} for row in cursor.fetchall()]
            return data

        elif func_name == "list_trains":
            keyword = func_args.get("keyword", "")
            limit = max(1, min(int(func_args.get("limit", 50) or 50), 200))
            page = max(1, int(func_args.get("page", 1) or 1))
            cursor = rw_conn.cursor()
            offset = (page - 1) * limit
            if keyword:
                cursor.execute(
                    "SELECT train_num FROM trains WHERE train_num LIKE ? ORDER BY train_num LIMIT ? OFFSET ?",
                    (f"%{keyword}%", limit, offset)
                )
            else:
                cursor.execute(
                    "SELECT train_num FROM trains ORDER BY train_num LIMIT ? OFFSET ?",
                    (limit, offset)
                )
            data = [{"train_num": row[0]} for row in cursor.fetchall()]
            return data

        elif func_name == "query_tickets":
            try:
                qid = question_id or get_current_question()
            except ValueError as e:
                return {"error": str(e)}

            db_path = get_question_db_path(qid)
            if not os.path.exists(db_path):
                return {"error": f"题目 {qid} 不存在"}

            q_conn = sqlite3.connect(db_path)
            try:
                seat_types = func_args.get("seat_types", None)
                from_station = func_args.get("from_station_id", None)
                to_station = func_args.get("to_station_id", None)
                tickets = get_train_tickets_filtered(
                    q_conn,
                    func_args.get("train_num", ""),
                    from_station,
                    to_station,
                    seat_types,
                )
                # 当指定了站对但数据库为空时，返回 0 张而非空对象
                if from_station and to_station:
                    types_to_check = seat_types if seat_types else TICKET_TABLES
                    key = f"{from_station}|{to_station}"
                    for t in types_to_check:
                        if t not in tickets:
                            tickets[t] = {}
                        if key not in tickets[t]:
                            tickets[t][key] = 0
                return tickets
            finally:
                q_conn.close()

        elif func_name == "query_ticket_price":
            from database import get_price, get_prices_conn
            p_conn = get_prices_conn()
            try:
                prices = {}
                for seat in ["class2", "class1", "class0"]:
                    p = get_price(
                        p_conn,
                        func_args.get("train_num", ""),
                        func_args.get("from_station_id", ""),
                        func_args.get("to_station_id", ""),
                        seat
                    )
                    if p is not None:
                        prices[seat] = p
            finally:
                p_conn.close()
            return prices if prices else None

        else:
            return {"error": f"未知工具: {func_name}"}

    except Exception as e:
        return {"error": str(e)}
    finally:
        rw_conn.close()


# ============================================================
# 第四部分：测评器接口
# ============================================================

class EvalCompleteRequest(BaseModel):
    question_id: str
    model_name: str
    user_input: str
    test_file: str
    score_summary: Dict
    verification: Dict
    human_confirmed: bool = False
    human_notes: str = ""


@app.post("/api/eval/complete")
def api_eval_complete(req: EvalCompleteRequest):
    """标记测评完成，保存结果"""
    from config import LOGS_RESULT_DIR

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{timestamp}_{req.model_name}_{req.question_id}.json"
    filepath = os.path.join(LOGS_RESULT_DIR, filename)

    record = {
        "timestamp": datetime.now().isoformat(),
        "test_file": req.test_file,
        "question_id": req.question_id,
        "model_name": req.model_name,
        "user_input": req.user_input,
        "score_summary": req.score_summary,
        "verification": req.verification,
        "human_confirmed": req.human_confirmed,
        "human_notes": req.human_notes,
        "status": "success",
    }

    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(record, f, ensure_ascii=False, indent=2)

    return {"success": True, "filepath": filepath, "message": f"测评结果已保存至 {filename}"}


@app.get("/api/test/records")
def api_test_records():
    """列出所有测试记录"""
    from config import LOGS_TEST_DIR
    records = []
    if os.path.exists(LOGS_TEST_DIR):
        for f in sorted(os.listdir(LOGS_TEST_DIR), reverse=True):
            if f.endswith(".json"):
                filepath = os.path.join(LOGS_TEST_DIR, f)
                try:
                    with open(filepath, "r", encoding="utf-8") as fh:
                        data = json.load(fh)
                    records.append({
                        "filename": f,
                        "timestamp": data.get("timestamp", ""),
                        "question_id": data.get("question_id", ""),
                        "model_name": data.get("model_name", ""),
                        "user_input": data.get("user_input", ""),
                    })
                except Exception:
                    records.append({"filename": f, "error": "读取失败"})
    return {"records": records, "total": len(records)}


class EvalVerifyRequest(BaseModel):
    question_id: str
    final_plan: List[Dict]


@app.post("/api/eval/verify")
def api_eval_verify(req: EvalVerifyRequest):
    """
    代码核查：对比最终乘车方案与余票数据库。
    - 从测试记录的 final_plan 中提取声称要购买的车次、座位、区间
    - 查对应题目的余票数据库，验证是否真的有票
    """
    if not req.question_id:
        raise HTTPException(status_code=400, detail="question_id 不能为空")
    if not req.final_plan:
        return {
            "total_items": 0,
            "correct_items": 0,
            "issue_count": 0,
            "hallucination_count": 0,
            "price_issue_count": 0,
            "verdict": "empty_plan",
            "results": [],
            "issues": [],
            "summary": "⚠️ 最终乘车方案为空（模型未输出 final_plan 或输出无解）",
        }

    try:
        result = verify_final_plan(req.final_plan, req.question_id)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"核查失败: {str(e)}")


@app.post("/api/eval/load")
def api_eval_load(filename: str = Body(..., embed=True)):
    """加载测试记录"""
    from config import LOGS_TEST_DIR
    filepath = os.path.join(LOGS_TEST_DIR, filename)
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail=f"测试记录 {filename} 不存在")
    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data


@app.get("/api/eval/results")
def api_eval_results():
    """列出所有测评结果"""
    from config import LOGS_RESULT_DIR
    results = []
    if os.path.exists(LOGS_RESULT_DIR):
        for f in sorted(os.listdir(LOGS_RESULT_DIR), reverse=True):
            if f.endswith(".json"):
                filepath = os.path.join(LOGS_RESULT_DIR, f)
                try:
                    with open(filepath, "r", encoding="utf-8") as fh:
                        data = json.load(fh)
                    results.append({
                        "filename": f,
                        "timestamp": data.get("timestamp", ""),
                        "question_id": data.get("question_id", ""),
                        "model_name": data.get("model_name", ""),
                        "score_summary": data.get("score_summary", {}),
                    })
                except Exception:
                    results.append({"filename": f, "error": "读取失败"})
    return {"results": results, "total": len(results)}


# ============================================================
# 第五部分：统计接口
# ============================================================

@app.get("/api/stats/summary")
def api_stats_summary():
    """获取统计汇总数据"""
    from statistics.aggregator import aggregate_results
    try:
        data = aggregate_results()
        return data
    except Exception as e:
        return {"error": str(e)}


@app.get("/api/stats/export/json")
def api_stats_export_json():
    """导出 JSON 格式报告"""
    from statistics.reporter import generate_report
    try:
        report = generate_report()
        return JSONResponse(content=report)
    except Exception as e:
        return {"error": str(e)}


@app.get("/api/stats/export/markdown")
def api_stats_export_markdown():
    """导出 Markdown 格式报告"""
    from statistics.reporter import export_markdown
    try:
        md = export_markdown()
        return HTMLResponse(content=md, media_type="text/markdown")
    except Exception as e:
        return {"error": str(e)}


# ============================================================
# 第五部分B：批量出题（1.xlsx 分布表 + 2.xlsx 站对表）
# ============================================================

# 分布行名称 → 题型参数映射（单策略 / 双策略 / 三策略；多策略一律“一次换乘类混合”）
_EXISTS_STRATEGY_MAP = {
    "换乘": {"question_type": "transfer", "transfers": 0, "segment_plans": []},
    "买短补长": {"question_type": "short_buy", "transfers": 0, "segment_plans": []},
    "额外购买（前）": {"question_type": "extra_front", "transfers": 0, "segment_plans": []},
    "额外购买（后）": {"question_type": "extra_rear", "transfers": 0, "segment_plans": []},
    # 双策略：前半段直达 + 后半段指定策略
    "买短补长+换乘": {"question_type": "mixed", "transfers": 1, "segment_plans": ["direct", "short_buy"]},
    "额外购买（前）+换乘": {"question_type": "mixed", "transfers": 1, "segment_plans": ["direct", "extra_front"]},
    "额外购买（后）+换乘": {"question_type": "mixed", "transfers": 1, "segment_plans": ["direct", "extra_rear"]},
    # 三策略：前后两段都有策略
    "买短补长+买短补长": {"question_type": "mixed", "transfers": 1, "segment_plans": ["short_buy", "short_buy"]},
    "买短补长+额外购买（前）": {"question_type": "mixed", "transfers": 1, "segment_plans": ["short_buy", "extra_front"]},
    "买短补长+额外购买（后）": {"question_type": "mixed", "transfers": 1, "segment_plans": ["short_buy", "extra_rear"]},
    "额外购买（前）+额外购买（前）": {"question_type": "mixed", "transfers": 1, "segment_plans": ["extra_front", "extra_front"]},
    "额外购买（前）+额外购买（后）": {"question_type": "mixed", "transfers": 1, "segment_plans": ["extra_front", "extra_rear"]},
    "额外购买（后）+额外购买（后）": {"question_type": "mixed", "transfers": 1, "segment_plans": ["extra_rear", "extra_rear"]},
}
_CRITERION_NAME_MAP = {
    "综合考虑": "comprehensive", "最快": "fastest", "最便宜": "cheapest",
    "出发最晚": "depart_latest", "最早到达": "arrive_earliest",
}
_CRITERION_CN = {v: k for k, v in _CRITERION_NAME_MAP.items()}
_BEHAVIOR_NAME_MAP = {
    "不允许换乘": "no_transfer",
    "不允许买短补长与额外购买": "no_short_buy_extra",
}
_BEHAVIOR_CN = {v: k for k, v in _BEHAVIOR_NAME_MAP.items()}
_BEHAVIOR_CN["none"] = "随意"


def _cell_int(v) -> int:
    """xlsx 单元格 → 非负整数（空值/文本返回 0）"""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return max(0, int(v))
    if isinstance(v, str) and v.strip().isdigit():
        return int(v.strip())
    return 0


def _parse_distribution_xlsx(data: bytes) -> Dict[str, Any]:
    """解析 1.xlsx 题目分布表（兼容原模板布局）。

    返回 {total: int|None, exists: [...], selective: [...]}，
    其中 exists 行含 category/name/has_interference/no_interference 及映射后的
    question_type/transfers/segment_plans；selective 行含 criterion/behavior/count。
    """
    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    wb.close()

    result: Dict[str, Any] = {"total": None, "exists": [], "selective": []}
    section: Optional[str] = None  # "exists" / "selective"
    current_category: Optional[str] = None
    current_criterion: Optional[str] = None

    for r in rows:
        cells = [str(c).strip() if c is not None else "" for c in r]
        joined = "".join(cells)
        if "总题目" in joined:
            for c in r:
                if isinstance(c, (int, float)) and not isinstance(c, bool):
                    result["total"] = int(c)
                    break
            continue
        if "存在性" in joined:
            section = "exists"
            current_category = None
            continue
        if "选择性" in joined:
            section = "selective"
            continue
        if not joined:
            continue

        if section == "exists":
            cat = cells[0] if cells[0] in ("单策略", "双策略", "三策略") else current_category
            name = cells[1] if len(cells) > 1 else ""
            config = _EXISTS_STRATEGY_MAP.get(name)
            if config:
                current_category = cat or "单策略"
                result["exists"].append({
                    "category": current_category or "单策略",
                    "name": name,
                    "has_interference": _cell_int(r[2]) if len(r) > 2 else 0,
                    "no_interference": _cell_int(r[3]) if len(r) > 3 else 0,
                    **{k: config[k] for k in ("question_type", "transfers", "segment_plans")},
                })
        elif section == "selective":
            crit = cells[0] if cells else ""
            if crit in _CRITERION_NAME_MAP:
                current_criterion = crit
            # 合并单元格：评判标准列为空时沿用上一行的标准
            if current_criterion is None:
                continue
            beh = cells[1] if len(cells) > 1 else "随意"
            cnt = _cell_int(r[2]) if len(r) > 2 else 0
            result["selective"].append({
                "criterion": _CRITERION_NAME_MAP[current_criterion],
                "behavior": _BEHAVIOR_NAME_MAP.get(beh, "none"),
                "count": cnt,
            })
    return result


def _parse_stations_xlsx(data: bytes) -> List[List[str]]:
    """解析 2.xlsx 到发站对：每行 [出发站, 到达站]，跳过空行与同站对。"""
    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    wb.close()
    stations: List[List[str]] = []
    for r in rows:
        if not r:
            continue
        a = str(r[0]).strip() if r[0] is not None else ""
        b = str(r[1]).strip() if len(r) > 1 and r[1] is not None else ""
        if a and b and a != b:
            stations.append([a, b])
    return stations


# 批量出题运行状态（供前端单一进度条轮询）
_batch_state: Dict[str, Any] = {
    "running": False,
    "done": False,
    "current": "",
    "total": 0,
    "done_count": 0,
    "result": None,
}


# ============================================================
# 批量工具日志：三个批量任务（batch_generate / batch_nl / batch_test）共用
# - 内存环形缓冲：供前端 status 接口增量轮询（seq 游标）
# - 追加落盘 logs/batch_tools.log：供后台直接查看
# ============================================================
_batch_logs: Dict[str, List[Dict[str, Any]]] = {
    "batch_generate": [], "batch_nl": [], "batch_test": [], "batch_eval": [],
}
_batch_logs_lock = threading.Lock()
_batch_log_seq = 0
_BATCH_LOG_MAX = 500                    # 每个任务内存中保留的最大条数
_BATCH_LOG_PATH = os.path.join(LOGS_DIR, "batch_tools.log")


def _batch_log(job: str, message: str, level: str = "info") -> None:
    """写一条批量任务日志：内存缓冲（带全局自增 seq）+ 追加落盘。"""
    global _batch_log_seq
    with _batch_logs_lock:
        _batch_log_seq += 1
        entry = {
            "seq": _batch_log_seq,
            "t": datetime.now().strftime("%H:%M:%S"),
            "level": level,
            "msg": message,
        }
        buf = _batch_logs.setdefault(job, [])
        buf.append(entry)
        if len(buf) > _BATCH_LOG_MAX:
            del buf[: len(buf) - _BATCH_LOG_MAX]
    try:
        os.makedirs(LOGS_DIR, exist_ok=True)
        with open(_BATCH_LOG_PATH, "a", encoding="utf-8") as f:
            f.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} [{job}] [{level}] {message}\n")
    except OSError:
        pass  # 落盘失败不影响任务执行


def _batch_log_clear(job: str) -> None:
    """任务启动时清空对应缓冲（全局 seq 不回退，前端游标安全）。"""
    with _batch_logs_lock:
        _batch_logs[job] = []


def _batch_log_payload(job: str, after_seq: int) -> Dict[str, Any]:
    """返回 seq > after_seq 的增量日志 + 当前全局 seq（前端下次带来的游标）。"""
    with _batch_logs_lock:
        items = [e for e in _batch_logs.get(job, []) if e["seq"] > after_seq]
        return {"items": items, "last_seq": _batch_log_seq}


@app.get("/api/batch/logs")
def api_batch_logs_file(lines: int = 500):
    """读取后台日志文件（logs/batch_tools.log）最后 N 行，纯文本输出，可直接在浏览器打开。"""
    if not os.path.exists(_BATCH_LOG_PATH):
        return Response(content="（日志文件尚不存在）", media_type="text/plain; charset=utf-8")
    try:
        with open(_BATCH_LOG_PATH, "r", encoding="utf-8") as f:
            all_lines = f.readlines()
        return Response(content="".join(all_lines[-max(1, lines):]), media_type="text/plain; charset=utf-8")
    except OSError as e:
        return Response(content=f"读取日志失败: {e}", media_type="text/plain; charset=utf-8")


def _generate_batch_nl(question_ids: List[str]) -> Dict[str, Any]:
    """批量落盘后为成功题目自动生成自然语言并写回 metadata['nl_question']。

    复用 nl_question.py 的 build_prompt / generate_nl（含评判标准/行为约束自然化）；
    未配置 NL_API_KEY 时整批跳过；单题失败不影响题目本身（只在结果中计入 failed）。
    返回 {"generated": n, "failed": n, "skipped": n, "error": str|None}。
    """
    from nl_question import build_prompt, generate_nl
    from config import ENV, ensure_env_fresh
    ensure_env_fresh()  # .env 被修改后自动重载
    api_key = (ENV.get("NL_API_KEY") or "").strip()
    if not api_key:
        return {"generated": 0, "failed": 0, "skipped": len(question_ids),
                "error": "未配置 NL_API_KEY，批量自然语言化已跳过（题目本身已成功落盘）"}
    model = (ENV.get("NL_MODEL") or ENV.get("DEFAULT_MODEL") or "").strip()
    base_url = (ENV.get("NL_BASE_URL") or ENV.get("DEFAULT_BASE_URL") or "").strip()
    if not model or not base_url:
        miss = "、".join(n for n, v in
                         [("NL_MODEL / DEFAULT_MODEL（模型名）", model),
                          ("NL_BASE_URL / DEFAULT_BASE_URL（接口地址）", base_url)] if not v)
        return {"generated": 0, "failed": 0, "skipped": len(question_ids),
                "error": f"未配置 {miss}，批量自然语言化已跳过（系统不默认指向任何平台；题目本身已成功落盘）"}
    metadata = load_metadata()
    generated = failed = 0
    for qid in question_ids:
        entry = metadata.get(qid)
        if not entry or not entry.get("question"):
            failed += 1
            continue
        try:
            nl = generate_nl(api_key, model, base_url, build_prompt(entry))
            entry["nl_question"] = nl
            generated += 1
        except Exception as e:
            failed += 1
    save_metadata(metadata)
    skipped = len(question_ids) - generated - failed
    return {"generated": generated, "failed": failed, "skipped": skipped, "error": None}


def _run_batch(payload: Dict[str, Any]) -> None:
    """在线程中顺序生成全部题目：每题固定参数（站对/方向/人数/座位/题型）重试 ≤ max_retries，
    成功即落盘（复用 _generate_one_variant + _confirm_generated_question），失败计入回执。"""
    rw_conn = get_railway_conn()
    try:
        distribution = payload.get("distribution") or []
        selective_rows = payload.get("selective") or []
        stations = payload.get("stations") or []
        seat_weights = payload.get("seat_weights") or {"class0": 20, "class1": 30, "class2": 50}
        # 干扰密度分开：伪干扰（存在性 1_ 有干扰）与真干扰（选择性 2_）；旧字段 interference_density 作为两者兜底
        fake_density = float(payload.get("fake_interference_density") or payload.get("interference_density") or 0.02)
        real_density = float(payload.get("real_interference_density") or payload.get("interference_density") or 0.02)
        max_retries = max(1, int(payload.get("max_retries") or 40))
        nl_enabled = bool(payload.get("nl_enabled", True))

        exists_total = sum((r.get("has_interference") or 0) + (r.get("no_interference") or 0) for r in distribution)
        sel_total = sum(r.get("count") or 0 for r in selective_rows)
        _batch_state.update({
            "running": True, "done": False, "current": "",
            "total": exists_total + sel_total,
            "done_count": 0,
            "result": None,
        })
        _batch_log("batch_generate",
                   f"批量出题开始：共 {exists_total + sel_total} 题（存在性 {exists_total} + 选择性 {sel_total}），每题重试上限 {max_retries}，伪干扰密度 {fake_density:.1%}，真干扰密度 {real_density:.1%}")

        details: List[Dict] = []
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        seq = 0

        def _build_request(question_type, transfers, segment_plans, mode, criterion, constraints, density):
            """随机取站对（方向可倒置）、随机人数 3~6、按权重随机座位等级；干扰密度按题型传入"""
            nonlocal seq
            pair = random.choice(stations)
            from_st, to_st = (pair[0], pair[1]) if random.random() < 0.5 else (pair[1], pair[0])
            seq += 1
            seat_type = random.choices(
                ["class0", "class1", "class2"],
                weights=[seat_weights.get("class0", 0), seat_weights.get("class1", 0), seat_weights.get("class2", 0)],
            )[0]
            return AutoGenerateRequest(
                question_type=question_type,
                from_station_id=from_st,
                to_station_id=to_st,
                mode=mode,
                random_tickets=False,
                fake_interference=False,
                interference_density=density,
                transfers=transfers,
                segment_plans=segment_plans,
                custom_qid=f"B{timestamp}_{seq:04d}",
                people_count=random.randint(3, 6),
                seat_type=seat_type,
                criterion=criterion,
                constraints=constraints,
            ), seq

        def _generate_one(question_type, transfers, plans, mode, criterion, constraints, prefix, fake, with_if, density):
            """重试 ≤ max_retries：每次重试都重新随机站对/方向/人数/座位（换站对提高成功率）；
            干扰密度按题型传入（存在性=伪干扰密度，选择性=真干扰密度）；
            成功返回 (qid, 尝试次数, None, seq)，失败返回 (None, max_retries, 错误, 首次seq)"""
            last_err = ""
            first_sq = None
            for attempt in range(1, max_retries + 1):
                req, sq = _build_request(question_type, transfers, plans, mode, criterion, constraints, density)
                if first_sq is None:
                    first_sq = sq
                try:
                    qid, _preview = _generate_one_variant(
                        req, prefix, fake,
                        "existence" if prefix != "2_" else "selective",
                        rw_conn, with_if,
                    )
                    _confirm_generated_question(qid, interference=with_if, interference_density=density)
                    return qid, attempt, None, sq
                except HTTPException as e:
                    last_err = str(e.detail) if hasattr(e, "detail") else str(e)
                except Exception as e:
                    last_err = str(e)
            return None, max_retries, last_err, first_sq

        # ---- 存在性问题：1_ 伪干扰（有干扰）/ 0_ 无干扰 ----
        exists_fail = [{"has_interference": 0, "no_interference": 0} for _ in distribution]
        for idx, row in enumerate(distribution):
            row_label = f"{row.get('category', '')}/{row.get('name', '')}"
            qtype = row.get("question_type", "transfer")
            transfers = int(row.get("transfers") or 0)
            plans = row.get("segment_plans") or []
            for _ in range(int(row.get("has_interference") or 0)):
                _batch_state["current"] = f"存在性（有干扰）：{row_label}"
                qid, attempt, err, sq = _generate_one(qtype, transfers, plans, "existence", "comprehensive", [], "1_", True, True, fake_density)
                details.append({
                    "question_id": qid or f"1_B{timestamp}_{sq:04d}",
                    "row": f"{row_label}/有干扰",
                    "attempts": attempt, "ok": qid is not None, "error": err,
                })
                if qid is None:
                    exists_fail[idx]["has_interference"] += 1
                    _batch_log("batch_generate", f"[{sq:04d}] 1_ 存在性（有干扰）{row_label} 失败（重试 {attempt} 次）：{err}", "error")
                else:
                    _batch_log("batch_generate", f"[{sq:04d}] 1_ 存在性（有干扰）{row_label} → {qid}（尝试 {attempt} 次）", "success")
                _batch_state["done_count"] += 1
            for _ in range(int(row.get("no_interference") or 0)):
                _batch_state["current"] = f"存在性（无干扰）：{row_label}"
                qid, attempt, err, sq = _generate_one(qtype, transfers, plans, "existence", "comprehensive", [], "0_", False, False, fake_density)
                details.append({
                    "question_id": qid or f"0_B{timestamp}_{sq:04d}",
                    "row": f"{row_label}/无干扰",
                    "attempts": attempt, "ok": qid is not None, "error": err,
                })
                if qid is None:
                    exists_fail[idx]["no_interference"] += 1
                    _batch_log("batch_generate", f"[{sq:04d}] 0_ 存在性（无干扰）{row_label} 失败（重试 {attempt} 次）：{err}", "error")
                else:
                    _batch_log("batch_generate", f"[{sq:04d}] 0_ 存在性（无干扰）{row_label} → {qid}（尝试 {attempt} 次）", "success")
                _batch_state["done_count"] += 1

        # ---- 选择性问题：2_ 真干扰（评判标准 + 行为约束） ----
        sel_fail = [0 for _ in selective_rows]
        for idx, row in enumerate(selective_rows):
            crit = row.get("criterion", "comprehensive")
            beh = row.get("behavior", "none")
            constraints = [] if beh == "none" else [beh]
            for _ in range(int(row.get("count") or 0)):
                _batch_state["current"] = f"选择性：{_CRITERION_CN.get(crit, crit)}({_BEHAVIOR_CN.get(beh, beh)})"
                # 题型不在此处指定：_generate_one_variant 对选择性题按行为约束自动推导（保证有解）
                qid, attempt, err, sq = _generate_one(None, 0, [], "selective", crit, constraints, "2_", False, True, real_density)
                details.append({
                    "question_id": qid or f"2_B{timestamp}_{sq:04d}",
                    "row": f"{_CRITERION_CN.get(crit, crit)}/{_BEHAVIOR_CN.get(beh, beh)}",
                    "attempts": attempt, "ok": qid is not None, "error": err,
                })
                if qid is None:
                    sel_fail[idx] += 1
                    _batch_log("batch_generate", f"[{sq:04d}] 2_ 选择性 {_CRITERION_CN.get(crit, crit)}({_BEHAVIOR_CN.get(beh, beh)}) 失败（重试 {attempt} 次）：{err}", "error")
                else:
                    _batch_log("batch_generate", f"[{sq:04d}] 2_ 选择性 {_CRITERION_CN.get(crit, crit)}({_BEHAVIOR_CN.get(beh, beh)}) → {qid}（尝试 {attempt} 次）", "success")
                _batch_state["done_count"] += 1

        ok_count = sum(1 for d in details if d["ok"])
        _batch_log("batch_generate",
                   f"出题完成：成功 {ok_count} / 失败 {len(details) - ok_count}（共 {len(details)} 题）",
                   "success" if ok_count == len(details) else "warn")
        nl_result = None
        if nl_enabled and ok_count > 0:
            nl_result = _generate_batch_nl([d["question_id"] for d in details if d["ok"]])
        report = {
            "total": len(details),
            "success": ok_count,
            "failed": len(details) - ok_count,
            "exists": [
                {"category": r.get("category", ""), "name": r.get("name", ""),
                 "has_interference_failed": exists_fail[i]["has_interference"],
                 "no_interference_failed": exists_fail[i]["no_interference"]}
                for i, r in enumerate(distribution)
            ],
            "selective": [
                {"criterion": r.get("criterion", ""), "behavior": r.get("behavior", ""),
                 "failed": sel_fail[i]}
                for i, r in enumerate(selective_rows)
            ],
        }
        _batch_state["result"] = {
            "summary": {"total": len(details), "success": ok_count, "failed": len(details) - ok_count,
                        "nl_generated": (nl_result or {}).get("generated", 0),
                        "nl_failed": (nl_result or {}).get("failed", 0),
                        "nl_skipped": (nl_result or {}).get("skipped", 0),
                        "nl_error": (nl_result or {}).get("error")},
            "details": details, "report": report,
        }
    except Exception as e:
        _batch_log("batch_generate", f"批量出题异常终止：{e}", "error")
        _batch_state["result"] = {
            "summary": {"total": 0, "success": 0, "failed": 0,
                        "nl_generated": 0, "nl_failed": 0, "nl_skipped": 0, "nl_error": str(e)},
            "details": [], "report": None, "error": str(e),
        }
    finally:
        rw_conn.close()
        _batch_state["running"] = False
        _batch_state["done"] = True


# --- POST /api/batch/parse-distribution 解析 1.xlsx 分布表 ---
@app.post("/api/batch/parse-distribution")
def api_batch_parse_distribution(file: UploadFile = File(...)):
    """上传 1.xlsx 题目分布表，解析为可编辑的结构化分布（前端展示并支持修改）。"""
    try:
        parsed = _parse_distribution_xlsx(file.file.read())
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"解析 1.xlsx 失败: {str(e)}")
    return {"success": True, **parsed}


# --- POST /api/batch/parse-stations 解析 2.xlsx 站对表 ---
@app.post("/api/batch/parse-stations")
def api_batch_parse_stations(file: UploadFile = File(...)):
    """上传 2.xlsx 到发站对表，返回站对列表（出题时随机选取、方向可倒置）。"""
    try:
        stations = _parse_stations_xlsx(file.file.read())
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"解析 2.xlsx 失败: {str(e)}")
    return {"success": True, "stations": stations}


class BatchGenerateRequest(BaseModel):
    distribution: List[Dict] = []
    selective: List[Dict] = []
    stations: List[List[str]] = []
    seat_weights: Dict[str, float] = {}
    interference_density: float = 0.02   # （兼容保留）未分字段时的兜底密度
    fake_interference_density: float = 0.02   # 伪干扰密度（存在性 1_ 有干扰）
    real_interference_density: float = 0.02   # 真干扰密度（选择性 2_）
    max_retries: int = 40
    # 批量落盘后是否自动生成自然语言并写回 metadata['nl_question']（需 .env NL_API_KEY）
    # 默认关闭：批量自然语言化由独立的批量语言框（/api/batch_nl/*）异步执行，与批量出题不混合
    nl_enabled: bool = False


# --- POST /api/batch/generate 启动批量出题（线程内执行，双进度条轮询） ---
@app.post("/api/batch/generate")
def api_batch_generate(req: BatchGenerateRequest):
    """按分布表 + 站对表一键批量出题：每题固定参数重试 ≤ max_retries，成功即直接落盘。"""
    if _batch_state.get("running"):
        raise HTTPException(status_code=400, detail="已有批量任务运行中，请等待完成")
    if not req.stations:
        raise HTTPException(status_code=400, detail="未提供站对表（2.xlsx），请先上传解析")
    import threading
    _batch_log_clear("batch_generate")
    _batch_log("batch_generate", "收到批量出题请求，任务已启动")
    thread = threading.Thread(target=_run_batch, args=(req.model_dump(),), daemon=True)
    thread.start()
    return {"success": True, "message": "批量出题已启动，可轮询 /api/batch/status 查看进度"}


# --- GET /api/batch/status 批量进度（单进度条 + 增量日志） ---
@app.get("/api/batch/status")
def api_batch_status(after_seq: int = 0):
    """返回批量出题进度：done_count/total 单一进度 + 当前任务 + 完成结果 + 增量日志。"""
    s = _batch_state
    return {
        "running": s["running"],
        "done": s["done"],
        "current": s["current"],
        "total": s["total"],
        "done_count": s["done_count"],
        "result": s["result"],
        "logs": _batch_log_payload("batch_generate", after_seq),
    }


# --- POST /api/batch/report 生成失败回执 xlsx（与 1.xlsx 同构） ---
@app.post("/api/batch/report")
def api_batch_report(report: Dict = Body(...)):
    """根据批量结果生成同构回执 xlsx：存在性区每格=该格失败数，选择性区每格=失败数，顶部=总失败数。"""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "批量回执"
    ws["A1"] = "总题目"
    ws["B1"] = report.get("total", 0)
    ws["A2"] = "成功"
    ws["B2"] = report.get("success", 0)
    ws["A3"] = "失败（总失败数）"
    ws["B3"] = report.get("failed", 0)

    ws["A5"] = "存在性问题"
    ws["C5"] = "有干扰失败"
    ws["D5"] = "无干扰失败"
    r = 6
    for item in report.get("exists", []):
        ws.cell(row=r, column=1, value=item.get("category", ""))
        ws.cell(row=r, column=2, value=item.get("name", ""))
        ws.cell(row=r, column=3, value=item.get("has_interference_failed", 0))
        ws.cell(row=r, column=4, value=item.get("no_interference_failed", 0))
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="选择性问题")
    ws.cell(row=r, column=3, value="失败数")
    r += 1
    for item in report.get("selective", []):
        ws.cell(row=r, column=1, value=_CRITERION_CN.get(item.get("criterion", ""), item.get("criterion", "")))
        ws.cell(row=r, column=2, value=_BEHAVIOR_CN.get(item.get("behavior", ""), item.get("behavior", "")))
        ws.cell(row=r, column=3, value=item.get("failed", 0))
        r += 1

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="batch_failure_report.xlsx"'},
    )


# ============================================================
# 第五部分C：批量自然语言化（独立框，与批量出题异步、互不影响）
# ============================================================

_nl_state: Dict[str, Any] = {
    "running": False, "done": False, "current": "",
    "total": 0, "done_count": 0, "result": None,
}


def _run_batch_nl(question_ids: List[str]) -> None:
    """在线程中对勾选题目逐个生成自然语言并写回 metadata['nl_question']。

    复用 nl_question.py 的 build_prompt（含评判标准/行为约束自然化）与 generate_nl；
    单题失败不影响其它题；缺 NL_API_KEY 时整批跳过。
    """
    from nl_question import build_prompt, generate_nl
    from config import ENV, ensure_env_fresh
    ensure_env_fresh()  # .env 被修改后自动重载
    api_key = (ENV.get("NL_API_KEY") or "").strip()
    if not api_key:
        _batch_log("batch_nl", "未配置 NL_API_KEY，批量自然语言化整批跳过", "error")
        _nl_state.update({
            "running": False, "done": True, "current": "",
            "result": {"summary": {"total": len(question_ids), "generated": 0, "failed": 0,
                                    "skipped": len(question_ids)},
                       "details": [], "error": "未配置 NL_API_KEY，批量自然语言化已整批跳过"},
        })
        return
    model = (ENV.get("NL_MODEL") or ENV.get("DEFAULT_MODEL") or "").strip()
    base_url = (ENV.get("NL_BASE_URL") or ENV.get("DEFAULT_BASE_URL") or "").strip()
    if not model or not base_url:
        miss = "、".join(n for n, v in
                         [("NL_MODEL / DEFAULT_MODEL（模型名）", model),
                          ("NL_BASE_URL / DEFAULT_BASE_URL（接口地址）", base_url)] if not v)
        _batch_log("batch_nl", f"未配置 {miss}，批量自然语言化整批跳过（系统不默认指向任何平台）", "error")
        _nl_state.update({
            "running": False, "done": True, "current": "",
            "result": {"summary": {"total": len(question_ids), "generated": 0, "failed": 0,
                                    "skipped": len(question_ids)},
                       "details": [], "error": f"未配置 {miss}，批量自然语言化已整批跳过"},
        })
        return
    _nl_state.update({"running": True, "done": False, "current": "",
                      "total": len(question_ids), "done_count": 0, "result": None})
    _batch_log("batch_nl", f"批量自然语言化开始：共 {len(question_ids)} 题（模型 {model}，温度 {NL_TEMPERATURE}）")
    metadata = load_metadata()
    details = []
    generated = failed = 0
    for i, qid in enumerate(question_ids, 1):
        entry = metadata.get(qid)
        _nl_state["current"] = f"自然语言化（{i}/{len(question_ids)}）：{qid}"
        _batch_log("batch_nl", f"（{i}/{len(question_ids)}）开始生成 {qid} 的自然语言")
        if not entry or not entry.get("question"):
            failed += 1
            details.append({"question_id": qid, "ok": False, "error": "题目缺少 question 字段"})
            _batch_log("batch_nl", f"{qid} 缺少 question 字段，跳过", "error")
        else:
            try:
                nl = generate_nl(api_key, model, base_url, build_prompt(entry))
                entry["nl_question"] = nl
                generated += 1
                details.append({"question_id": qid, "ok": True, "error": None})
                # 每题成功立即落盘：进度实时可见，中断不丢已生成结果
                save_metadata(metadata)
                _batch_log("batch_nl", f"{qid} 自然语言生成成功并已写回 metadata", "success")
            except Exception as e:
                failed += 1
                details.append({"question_id": qid, "ok": False, "error": str(e)})
                _batch_log("batch_nl", f"{qid} 自然语言生成失败：{e}", "error")
        _nl_state["done_count"] += 1
    save_metadata(metadata)
    _nl_state["result"] = {
        "summary": {"total": len(question_ids), "generated": generated, "failed": failed,
                    "skipped": len(question_ids) - generated - failed},
        "details": details, "error": None,
    }
    _batch_log("batch_nl", f"批量自然语言化完成：成功 {generated} / 失败 {failed} / 跳过 {len(question_ids) - generated - failed}",
               "success" if failed == 0 else "warn")
    _nl_state.update({"running": False, "done": True, "current": ""})


class BatchNlGenerateRequest(BaseModel):
    question_ids: List[str] = []


@app.get("/api/batch_nl/scan")
def api_batch_nl_scan():
    """扫描缺失自然语言（nl_question）的题目列表，供批量语言面板勾选。"""
    metadata = load_metadata()
    from config import QUESTION_DIR
    db_names = {f[:-3] for f in os.listdir(QUESTION_DIR) if f.endswith(".db")} if os.path.isdir(QUESTION_DIR) else set()
    items = []
    for qid, m in metadata.items():
        if not isinstance(m, dict):
            continue
        if m.get("nl_question"):
            continue
        items.append({
            "question_id": qid,
            "type": m.get("type", ""),
            "question_type": m.get("question_type", ""),
            "question": m.get("question", ""),
            "people_count": m.get("people_count"),
            "seat_type": m.get("seat_type"),
            "criterion": m.get("criterion"),
            "constraints": m.get("constraints"),
            "db_exists": qid in db_names,
        })
    return {"success": True, "items": items, "total": len(items)}


@app.post("/api/batch_nl/generate")
def api_batch_nl_generate(req: BatchNlGenerateRequest):
    """启动批量自然语言化（独立线程与进度，与批量出题互不影响）。"""
    if _nl_state.get("running"):
        raise HTTPException(status_code=400, detail="已有批量自然语言任务运行中，请等待完成")
    if not req.question_ids:
        raise HTTPException(status_code=400, detail="未选择任何题目（可先扫描缺失自然语言的题目）")
    import threading
    _batch_log_clear("batch_nl")
    _batch_log("batch_nl", f"收到批量自然语言化请求：{len(req.question_ids)} 题，任务已启动")
    threading.Thread(target=_run_batch_nl, args=(list(req.question_ids),), daemon=True).start()
    return {"success": True, "message": "批量自然语言化已启动（逐题生成并写回 nl_question）"}


@app.get("/api/batch_nl/status")
def api_batch_nl_status(after_seq: int = 0):
    """返回批量自然语言化进度（单进度条）+ 增量日志 + 结果。"""
    s = _nl_state
    return {"running": s["running"], "done": s["done"], "current": s["current"],
            "total": s["total"], "done_count": s["done_count"], "result": s["result"],
            "logs": _batch_log_payload("batch_nl", after_seq)}


# ============================================================
# 第五部分D：批量测试（扫描可测题目 → 勾选 → 逐题测试）
# ============================================================

_test_state: Dict[str, Any] = {
    "running": False, "done": False, "current": "",
    "total": 0, "done_count": 0, "result": None,
}


def _batch_test_model_testable(m: Dict, model: str) -> bool:
    """可测试条件：题目信息核查完备（含自然语言问题）且该模型未测试过。"""
    return bool(m.get("question")) and bool(m.get("nl_question")) \
        and model not in (m.get("tested_models") or [])


def _run_batch_test(model: str, question_ids: List[str], max_iterations: int, concurrency: int = 1) -> None:
    """线程池并发测试：每题独立会话（session=batch_test_<qid>），题目上下文按调用传入（不依赖全局当前题目）；
    成功后把模型名写入该题 metadata.tested_models（每个模型每题只测一次），最后统一落盘 metadata。"""
    concurrency = max(1, min(int(concurrency or 1), 8))
    _test_state.update({"running": True, "done": False, "current": "",
                        "total": len(question_ids), "done_count": 0, "result": None})
    _batch_log("batch_test", f"批量测试开始：模型 {model}，共 {len(question_ids)} 题，最大工具调用 {max_iterations}，并发数 {concurrency}")
    metadata = load_metadata()
    details = []
    state_lock = threading.Lock()

    def _test_one(seq: int, qid: str) -> Dict:
        """测试单题（worker 线程内执行），返回 detail 字典。"""
        entry = metadata.get(qid) or {}
        _test_state["current"] = f"并发 {concurrency} · 已完成 {_test_state['done_count']}/{len(question_ids)} · 最近开始：{qid}"
        _batch_log("batch_test", f"（{seq}/{len(question_ids)}）开始测试 {qid}")
        if not _batch_test_model_testable(entry, model):
            _batch_log("batch_test", f"{qid} 不可测（缺 nl_question 或该模型已测过），跳过", "warn")
            return {"question_id": qid, "ok": False,
                    "error": "题目信息不完备（缺 nl_question）或该模型已测试过，已跳过（去重）"}
        try:
            session_id = f"batch_test_{qid}"
            api_test_reset(TestResetRequest(session_id=session_id))
            chat_r = asyncio.run(api_test_chat(ChatRequest(
                message=entry.get("nl_question", ""),
                model_name=model,
                question_id=qid,
                session_id=session_id,
                max_iterations=max_iterations,
            )))
            if chat_r.get("error"):
                raise RuntimeError(str(chat_r["error"]))
            complete_r = api_test_complete(TestCompleteRequest(session_id=session_id))
            if not complete_r.get("success"):
                raise RuntimeError(str(complete_r.get("error", "测试记录保存失败")))
            tested = list(entry.get("tested_models") or [])
            if model not in tested:
                tested.append(model)
            entry["tested_models"] = tested
            _batch_log("batch_test",
                       f"{qid} 测试完成（plan_status={complete_r.get('plan_status', '')}，耗时 {round(complete_r.get('duration', 0), 1)}s）→ {complete_r.get('filename', '')}",
                       "success")
            return {
                "question_id": qid, "ok": True, "error": None,
                "filename": complete_r.get("filename", ""),
                "plan_status": complete_r.get("plan_status", ""),
                "duration": round(complete_r.get("duration", 0), 1),
            }
        except Exception as e:
            _batch_log("batch_test", f"{qid} 测试失败：{e}", "error")
            return {"question_id": qid, "ok": False, "error": str(e)}

    from concurrent.futures import ThreadPoolExecutor, as_completed
    with ThreadPoolExecutor(max_workers=concurrency) as pool:
        futures = [pool.submit(_test_one, i, qid) for i, qid in enumerate(question_ids, 1)]
        for fut in as_completed(futures):
            detail = fut.result()
            with state_lock:
                details.append(detail)
                _test_state["done_count"] += 1
    save_metadata(metadata)
    ok_count = sum(1 for d in details if d["ok"])
    details.sort(key=lambda d: question_ids.index(d["question_id"]))
    _test_state["result"] = {
        "summary": {"total": len(question_ids), "success": ok_count,
                    "failed": len(details) - ok_count, "model": model,
                    "concurrency": concurrency},
        "details": details,
    }
    _batch_log("batch_test", f"批量测试完成：成功 {ok_count} / 失败 {len(details) - ok_count}（共 {len(question_ids)} 题）",
               "success" if ok_count == len(details) else "warn")
    _test_state.update({"running": False, "done": True, "current": ""})


class BatchTestScanRequest(BaseModel):
    model: str = ""


class BatchTestStartRequest(BaseModel):
    model: str = ""
    question_ids: List[str] = []
    max_iterations: int = 30
    concurrency: int = 1


@app.post("/api/batch_test/scan")
def api_batch_test_scan(req: BatchTestScanRequest):
    """扫描可测试题目：信息核查完备（含 nl_question）且该模型未测过；前端据此勾选增删。"""
    model = (req.model or "").strip()
    if not model:
        raise HTTPException(status_code=400, detail="请先填写测试模型编号（名称）")
    metadata = load_metadata()
    from config import QUESTION_DIR
    db_names = {f[:-3] for f in os.listdir(QUESTION_DIR) if f.endswith(".db")} if os.path.isdir(QUESTION_DIR) else set()
    items = []
    for qid, m in metadata.items():
        if not isinstance(m, dict):
            continue
        if qid not in db_names:
            continue
        items.append({
            "question_id": qid,
            "type": m.get("type", ""),
            "question_type": m.get("question_type", ""),
            "question": m.get("question", ""),
            "nl_exists": bool(m.get("nl_question")),
            "people_count": m.get("people_count"),
            "seat_type": m.get("seat_type"),
            "criterion": m.get("criterion"),
            "constraints": m.get("constraints"),
            "tested_models": m.get("tested_models") or [],
            "testable": _batch_test_model_testable(m, model),
        })
    return {"success": True, "model": model, "items": items, "total": len(items)}


@app.post("/api/batch_test/start")
def api_batch_test_start(req: BatchTestStartRequest):
    """启动批量测试：仅对“信息完备且该模型未测过”的题目执行，逐题保存测试记录并写回 tested_models。"""
    model = (req.model or "").strip()
    if not model:
        raise HTTPException(status_code=400, detail="请先填写测试模型编号（名称）")
    if _test_state.get("running"):
        raise HTTPException(status_code=400, detail="已有批量测试任务运行中，请等待完成")
    if not req.question_ids:
        raise HTTPException(status_code=400, detail="未选择任何题目（可先扫描可测试题目）")
    import threading
    max_iter = max(1, min(int(req.max_iterations or 30), 100))
    concurrency = max(1, min(int(req.concurrency or 1), 8))
    _batch_log_clear("batch_test")
    _batch_log("batch_test", f"收到批量测试请求：模型 {model}，{len(req.question_ids)} 题，并发 {concurrency}，任务已启动")
    threading.Thread(target=_run_batch_test, args=(model, list(req.question_ids), max_iter, concurrency), daemon=True).start()
    return {"success": True, "message": f"批量测试已启动（并发 {concurrency}，仅对信息完备且该模型未测过的题目执行）"}


@app.get("/api/batch_test/status")
def api_batch_test_status(after_seq: int = 0):
    """返回批量测试进度（单进度条）+ 增量日志 + 逐题结果。"""
    s = _test_state
    return {"running": s["running"], "done": s["done"], "current": s["current"],
            "total": s["total"], "done_count": s["done_count"], "result": s["result"],
            "logs": _batch_log_payload("batch_test", after_seq)}


# ============================================================
# 第五部分E：批量测评（勾选测试记录 → 批量代码核查并保存结果）
# ============================================================

_eval_state: Dict[str, Any] = {
    "running": False, "done": False, "current": "",
    "total": 0, "done_count": 0, "result": None,
}


def _list_evaluated_test_files() -> set:
    """已测评的测试记录文件名集合（来自 logs/result 的 test_file 字段）。"""
    from config import LOGS_RESULT_DIR
    evaluated = set()
    if os.path.exists(LOGS_RESULT_DIR):
        for f in os.listdir(LOGS_RESULT_DIR):
            if not f.endswith(".json"):
                continue
            try:
                with open(os.path.join(LOGS_RESULT_DIR, f), "r", encoding="utf-8") as fh:
                    data = json.load(fh)
                tf = data.get("test_file")
                if tf:
                    evaluated.add(tf)
            except Exception:
                continue
    return evaluated


def _run_batch_eval(filenames: List[str]) -> None:
    """串行逐条测评：读测试记录 → 代码核查（verifier）→ 保存结果到 logs/result（结构同手动测评）。"""
    from config import LOGS_TEST_DIR, LOGS_RESULT_DIR
    _eval_state.update({"running": True, "done": False, "current": "",
                        "total": len(filenames), "done_count": 0, "result": None})
    _batch_log("batch_eval", f"批量测评开始：共 {len(filenames)} 条测试记录")
    details = []
    ok_count = 0
    verdict_count: Dict[str, int] = {}
    for i, fn in enumerate(filenames, 1):
        _eval_state["current"] = f"测评（{i}/{len(filenames)}）：{fn}"
        _batch_log("batch_eval", f"（{i}/{len(filenames)}）开始核查 {fn}")
        filepath = os.path.join(LOGS_TEST_DIR, fn)
        try:
            if not os.path.exists(filepath):
                raise FileNotFoundError(f"测试记录 {fn} 不存在")
            with open(filepath, "r", encoding="utf-8") as fh:
                data = json.load(fh)
            qid = data.get("question_id", "") or "unknown"
            final_plan = data.get("final_plan") or []
            verification = verify_final_plan(final_plan, qid)
            verdict = verification.get("verdict", "unknown")
            verdict_count[verdict] = verdict_count.get(verdict, 0) + 1
            tu = data.get("token_usage") or {}
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            result_name = f"{timestamp}_{data.get('model_name', 'unknown')}_{qid}.json"
            record = {
                "timestamp": datetime.now().isoformat(),
                "test_file": fn,
                "question_id": qid,
                "model_name": data.get("model_name", "unknown"),
                "user_input": data.get("user_input", ""),
                "score_summary": {
                    "hallucination_count": verification.get("hallucination_count", 0),
                    "issue_count": verification.get("issue_count", 0),
                    "total_tokens": tu.get("total_tokens", 0),
                    "duration_seconds": data.get("duration", 0),
                    "tool_calls": [],
                },
                "verification": verification,
                "human_confirmed": False,
                "human_notes": "批量测评自动生成",
                "status": "success",
            }
            with open(os.path.join(LOGS_RESULT_DIR, result_name), "w", encoding="utf-8") as fh:
                json.dump(record, fh, ensure_ascii=False, indent=2)
            ok_count += 1
            details.append({"filename": fn, "question_id": qid, "ok": True, "error": None,
                            "verdict": verdict, "result_file": result_name})
            _batch_log("batch_eval", f"{fn} → {qid} 核查完成：verdict={verdict}，结果已保存 {result_name}", "success")
        except Exception as e:
            verdict_count["error"] = verdict_count.get("error", 0) + 1
            details.append({"filename": fn, "question_id": "", "ok": False, "error": str(e),
                            "verdict": "error", "result_file": ""})
            _batch_log("batch_eval", f"{fn} 测评失败：{e}", "error")
        _eval_state["done_count"] += 1
    _eval_state["result"] = {
        "summary": {"total": len(filenames), "success": ok_count,
                    "failed": len(details) - ok_count, "verdicts": verdict_count},
        "details": details,
    }
    _batch_log("batch_eval", f"批量测评完成：成功 {ok_count} / 失败 {len(details) - ok_count}，verdict 分布 {verdict_count}",
               "success" if ok_count == len(details) else "warn")
    _eval_state.update({"running": False, "done": True, "current": ""})


@app.post("/api/batch_eval/scan")
def api_batch_eval_scan():
    """扫描测试记录列表（含题号/模型/时间/plan_status/是否已测评），供批量测评勾选。"""
    from config import LOGS_TEST_DIR
    evaluated = _list_evaluated_test_files()
    items = []
    if os.path.exists(LOGS_TEST_DIR):
        for f in sorted(os.listdir(LOGS_TEST_DIR), reverse=True):
            if not f.endswith(".json"):
                continue
            filepath = os.path.join(LOGS_TEST_DIR, f)
            try:
                with open(filepath, "r", encoding="utf-8") as fh:
                    data = json.load(fh)
                items.append({
                    "filename": f,
                    "timestamp": data.get("timestamp", ""),
                    "question_id": data.get("question_id", ""),
                    "model_name": data.get("model_name", ""),
                    "plan_status": data.get("plan_status", ""),
                    "duration": data.get("duration", 0),
                    "evaluated": f in evaluated,
                })
            except Exception:
                items.append({"filename": f, "error": "读取失败", "evaluated": False})
    return {"success": True, "items": items, "total": len(items)}


@app.post("/api/batch_eval/start")
def api_batch_eval_start(req: Dict = Body(...)):
    """启动批量测评：对勾选的测试记录逐条代码核查并保存结果。"""
    filenames = [str(f) for f in (req.get("records") or []) if f]
    if _eval_state.get("running"):
        raise HTTPException(status_code=400, detail="已有批量测评任务运行中，请等待完成")
    if not filenames:
        raise HTTPException(status_code=400, detail="未选择任何测试记录（可先扫描测评记录）")
    _batch_log_clear("batch_eval")
    _batch_log("batch_eval", f"收到批量测评请求：{len(filenames)} 条记录，任务已启动")
    threading.Thread(target=_run_batch_eval, args=(filenames,), daemon=True).start()
    return {"success": True, "message": "批量测评已启动（逐条代码核查并保存结果）"}


@app.get("/api/batch_eval/status")
def api_batch_eval_status(after_seq: int = 0):
    """返回批量测评进度（单进度条）+ 增量日志 + 结果。"""
    s = _eval_state
    return {"running": s["running"], "done": s["done"], "current": s["current"],
            "total": s["total"], "done_count": s["done_count"], "result": s["result"],
            "logs": _batch_log_payload("batch_eval", after_seq)}


# ============================================================
# 第六部分：静态文件服务
# ============================================================

# 挂载静态文件目录
STATIC_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static")
if os.path.exists(STATIC_DIR):
    app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


# 前端页面路由（SPA：所有页面路由返回 index.html，JS 负责显示对应页面）


@app.get("/", response_class=HTMLResponse)
def index():
    return get_html("index.html")


@app.get("/manual_question", response_class=HTMLResponse)
def manual_question():
    return get_html("index.html")


@app.get("/auto_question", response_class=HTMLResponse)
def auto_question():
    return get_html("index.html")


@app.get("/selective_question", response_class=HTMLResponse)
def selective_question():
    return get_html("index.html")


@app.get("/batch_question", response_class=HTMLResponse)
def batch_question():
    return get_html("index.html")


@app.get("/batch_nl_question", response_class=HTMLResponse)
def batch_nl_question():
    return get_html("index.html")


@app.get("/batch_test_question", response_class=HTMLResponse)
def batch_test_question():
    return get_html("index.html")


@app.get("/question_manager", response_class=HTMLResponse)
def question_manager():
    return get_html("index.html")


@app.get("/edit_question", response_class=HTMLResponse)
def edit_question():
    return get_html("index.html")


@app.get("/stats", response_class=HTMLResponse)
def stats():
    return get_html("index.html")


@app.get("/eval", response_class=HTMLResponse)
def eval_page():
    return get_html("index.html")


def get_html(filename: str) -> str:
    """读取 HTML 文件"""
    filepath = os.path.join(TEMPLATES_DIR, filename)
    if os.path.exists(filepath):
        with open(filepath, "r", encoding="utf-8") as f:
            return f.read()
    return f"<html><body><h1>404</h1><p>{filename} 未找到</p></body></html>"


# ============================================================
# 启动入口
# ============================================================
if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host=API_CONFIG["host"], port=API_CONFIG["port"])