"""
assist/station_stop_counts.py
统计每个车站的停车数（被多少个车次停靠）并导出 Excel。

用法（命令行终端）：
    python assist/station_stop_counts.py [-o 输出.xlsx] [--db 基础数据库路径]

- 数据来源：data/railway.db 的 train_stops 表（默认路径，可通过 --db 覆盖）
- 输出 Excel 列：
    1. 站名
    2. 站ID（电报码）
    3. 停车数（停靠该站的不同车次数，DISTINCT train_num）
    4. 停靠总次数（含车次重复出现的经停记录数）
- 默认按停车数降序排列
- 未指定 -o 时，输出文件放在本脚本所在目录（assist/）：
    assist/station_stop_counts.xlsx
"""

import argparse
import os
import sys

import openpyxl

# 项目根目录（assist 与 docs 同级，项目根在上一级）
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)

from config import RAILWAY_DB_PATH  # noqa: E402
from database import get_railway_conn  # noqa: E402


def collect_stop_counts(conn) -> list:
    """统计每个车站的停车数，返回 [(station_id, station_name, distinct_trains, total_records), ...]"""
    cursor = conn.cursor()
    cursor.execute("""
        SELECT
            ts.station_id,
            MAX(s.station_name) AS station_name,
            COUNT(DISTINCT ts.train_num) AS distinct_trains,
            COUNT(*) AS total_records
        FROM train_stops ts
        LEFT JOIN stations s ON s.station_id = ts.station_id
        GROUP BY ts.station_id
        ORDER BY distinct_trains DESC, station_name ASC
    """)
    return [tuple(row) for row in cursor.fetchall()]


def write_xlsx(rows: list, out_path: str) -> None:
    """将统计结果写入 Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "每站停车数"
    ws.append(["站名", "站ID（电报码）", "停车数（不同车次数）", "停靠总次数（含重复经停）"])
    for station_id, station_name, distinct_trains, total_records in rows:
        ws.append([station_name or station_id, station_id, distinct_trains, total_records])
    # 表头加粗
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    # 列宽
    for col_idx, width in [(1, 18), (2, 12), (3, 20), (4, 24)]:
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    wb.save(out_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="统计每个车站的停车数并导出 Excel")
    parser.add_argument("-o", "--output", default=None,
                        help="输出 Excel 路径（默认 = assist/station_stop_counts.xlsx，即本脚本所在目录）")
    parser.add_argument("--db", default=None, help="基础数据库路径（默认 data/railway.db）")
    args = parser.parse_args()

    db_path = args.db or RAILWAY_DB_PATH
    out_path = args.output or os.path.join(os.path.dirname(os.path.abspath(__file__)), "station_stop_counts.xlsx")

    if not os.path.exists(db_path):
        print(f"[错误] 基础数据库不存在: {db_path}，请先运行 collector.py 采集数据", file=sys.stderr)
        return 1

    import sqlite3
    conn = sqlite3.connect(db_path)
    try:
        rows = collect_stop_counts(conn)
        if not rows:
            print("[警告] 数据库中没有经停记录（train_stops 为空）", file=sys.stderr)
        write_xlsx(rows, out_path)
        print(f"[完成] 共统计 {len(rows)} 个车站，已导出: {out_path}")
    finally:
        conn.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())