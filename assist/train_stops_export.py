"""
assist/train_stops_export.py
车次停站序列导出：输入一个 Excel（第一列为一竖列车号），
输出一个 Excel，第一列保留车号，后面跟着该车次的停站序列（站名按 stop_no 升序逐列展开）。

用法（命令行终端）：
    python assist/train_stops_export.py 输入车号表.xlsx [-o 输出.xlsx] [--db 基础数据库路径]

- 输入 Excel：第一列每一行一个车次号（如 G1、G2…），多余列忽略
- 输出 Excel 列：
    1. 车次号（原文）
    2. 经停站数
    3. 停站1（第1站站名）
    4. 停站2（第2站站名）
    … 依次到最后一站；缺失的车次在「备注」列标注「车次不存在」
    最后一列 = 备注
- 未指定 -o 时，输出文件放在本脚本所在目录（assist/）：
    assist/train_stops_export_out.xlsx
"""

import argparse
import os
import sqlite3
import sys

import openpyxl

# 项目根目录（assist 与 docs 同级，项目根在上一级）
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)

from config import RAILWAY_DB_PATH  # noqa: E402


def read_train_nums(path: str) -> list:
    """读取输入 Excel 第一列的车次号（跳过空行）"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    nums = []
    for row in wb.active.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        v = str(row[0]).strip()
        if v:
            nums.append(v)
    wb.close()
    return nums


def get_stops(conn: sqlite3.Connection, train_num: str) -> list:
    """按 stop_no 升序返回该车次停站名列表"""
    cursor = conn.cursor()
    cursor.execute("""
        SELECT station_name FROM train_stops
        WHERE train_num = ?
        ORDER BY stop_no
    """, (train_num,))
    return [row[0] for row in cursor.fetchall()]


def write_result(nums: list, rows: list, out_path: str) -> None:
    """输出：车次号 + 经停站数 + 逐列站名 + 备注"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "车次停站序列"
    max_stops = max((len(r[1]) for r in rows), default=0)
    header = ["车次号", "经停站数"] + [f"停站{i}" for i in range(1, max_stops + 1)] + ["备注"]
    ws.append(header)
    for num, stops, note in rows:
        ws.append([num, len(stops)] + stops + [""] * (max_stops - len(stops)) + [note])
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    for col_idx, width in [(1, 12), (2, 10)]:
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    for i in range(3, 3 + max_stops):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 14
    ws.column_dimensions[openpyxl.utils.get_column_letter(3 + max_stops)].width = 18
    wb.save(out_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="导出车次停站序列到 Excel")
    parser.add_argument("input", help="输入车号表 Excel（第一列为一竖列车号）")
    parser.add_argument("-o", "--output", default=None,
                        help="输出 Excel 路径（默认 = assist/train_stops_export_out.xlsx）")
    parser.add_argument("--db", default=None, help="基础数据库路径（默认 data/railway.db）")
    args = parser.parse_args()

    db_path = args.db or RAILWAY_DB_PATH
    out_path = args.output or os.path.join(os.path.dirname(os.path.abspath(__file__)), "train_stops_export_out.xlsx")

    if not os.path.exists(db_path):
        print(f"[错误] 基础数据库不存在: {db_path}，请先运行 collector.py 采集数据", file=sys.stderr)
        return 1
    if not os.path.exists(args.input):
        print(f"[错误] 输入文件不存在: {args.input}", file=sys.stderr)
        return 1

    conn = sqlite3.connect(db_path)
    try:
        nums = read_train_nums(args.input)
        if not nums:
            print("[警告] 输入中没有有效的车次号", file=sys.stderr)
        rows = []
        for num in nums:
            stops = get_stops(conn, num)
            if not stops:
                rows.append([num, [], "车次不存在"])
            else:
                rows.append([num, stops, ""])
        write_result(nums, rows, out_path)
        print(f"[完成] 共导出 {len(nums)} 个车次，已导出: {out_path}")
    finally:
        conn.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())