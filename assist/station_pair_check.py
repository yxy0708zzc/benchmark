"""
assist/station_pair_check.py
站对检查：输入一个 Excel（站对表，格式同 2.xlsx：每行 [出发站, 到达站]），
输出一个 Excel，在第三、四列展示：
    3. 是否有直达车（有直达 → 是；无 → 否）
    4. 是否能通过换乘实现（一次换乘可行 → 是；否则 → 否）

换乘判定与自动出题完全一致（同 server.py 的算法）：
- 候选首程车 T：经过出发站 A、不经过到达站 B、A 之后还有中间站
- 换乘衔接最短时间 min_gap = 20 分钟（与自动出题固定一致，可用 --min-gap 调整）
- 枚举 T 的全部中间站 M，找二程车 U：U 经过 M 且在 M 之后经过 B、U ≠ T、
  且 T 到达 M 的时刻 + min_gap ≤ U 从 M 出发的时刻
- 重试次数与自动出题相同：外层对候选 T 至多重试 min(len(T候选), 15) 辆
  （可用 --max-attempts 调整），任一成功即判「能换乘」
- 中间站过滤：与到达站 B 同城的中间站被排除（与自动出题一致）

用法（命令行终端）：
    python assist/station_pair_check.py 输入站对表.xlsx [-o 输出.xlsx] [--db 基础数据库路径] [--min-gap 20] [--max-attempts 15]

- 输入 Excel 每行两列：出发站、到达站（站名或电报码均可），空行与同站对自动跳过
- 输出 Excel 列：
    1. 出发站（原文）
    2. 到达站（原文）
    3. 是否有直达车（是/否）
    4. 是否能通过换乘实现（是/否）
    5. 直达车次列表（有直达时，顿号分隔；否则留空）
    6. 示例换乘方案（能换乘时为「T 车次 A→M 换 U 车次 M→B」；否则留空）
    7. 备注（站不存在等说明）
- 未指定 -o 时，输出文件放在本脚本所在目录（assist/）：
    assist/station_pair_check_out.xlsx
"""

import argparse
import os
import random
import sqlite3
import sys

import openpyxl

# 项目根目录（assist 与 docs 同级，项目根在上一级）
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)

from config import RAILWAY_DB_PATH  # noqa: E402
from database import get_train_stops, get_routes_between, resolve_station_name_or_id  # noqa: E402


# ---------------------------------------------------------------------------
# 与自动出题一致的换乘判定算法（复制 server.py 的核心逻辑）
# ---------------------------------------------------------------------------

def _get_city_key(station_name: str) -> str:
    """从站名提取城市标识（用于判断同城车站，与自动出题一致）"""
    for suf in ['东', '南', '西', '北', '站']:
        if station_name.endswith(suf) and len(station_name) > len(suf) + 1:
            return station_name[:-len(suf)]
    return station_name


def _calc_time_diff_minutes(time1: str, time2: str) -> int:
    """计算 time2 - time1 的分钟数（仅当日；跨天返回负值），与自动出题一致"""
    try:
        h1, m1 = map(int, time1.split(":"))
        h2, m2 = map(int, time2.split(":"))
        return (h2 * 60 + m2) - (h1 * 60 + m1)
    except (ValueError, AttributeError):
        return 0


def _train_passes_station(conn: sqlite3.Connection, train_num: str, station_id: str) -> bool:
    cursor = conn.cursor()
    cursor.execute("SELECT 1 FROM train_stops WHERE train_num = ? AND station_id = ?",
                   (train_num, station_id))
    return cursor.fetchone() is not None


def _get_trains_passing_station(conn: sqlite3.Connection, station_id: str):
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT train_num FROM train_stops WHERE station_id = ?", (station_id,))
    return [row[0] for row in cursor.fetchall()]


def _get_routes_for_transfer(conn: sqlite3.Connection, from_id: str, to_id: str):
    """候选首程车 T：经过 A、不经过 B、A 之后至少还有一站（与自动出题一致）"""
    routes = []
    for tn in _get_trains_passing_station(conn, from_id):
        if _train_passes_station(conn, tn, to_id):
            continue  # 排除经过 B 的车次（防止直达逃逸）
        stops = get_train_stops(conn, tn)
        stop_ids = [s["station_id"] for s in stops]
        if from_id not in stop_ids:
            continue
        from_idx = stop_ids.index(from_id)
        if from_idx < len(stops) - 1:
            routes.append({
                "train_num": tn,
                "depart_time": stops[from_idx]["stop_time"],
                "arrive_time": stops[-1]["stop_time"],
                "stops": stops,
                "from_idx": from_idx,
            })
    return routes


def _find_transfer_solutions(conn: sqlite3.Connection,
                             excluded_train_num: str,
                             mid_refs,
                             to_station_id: str,
                             min_gap: int = 20):
    """一次枚举全部可行换乘方案 (M, U)：U 经过 M 且在 M 之后经过 B、U ≠ T、
    参考到达时间 + min_gap ≤ U 从 M 出发时间。（与自动出题 _find_transfer_solutions 一致）"""
    if not mid_refs:
        return []
    mid_ids = [m for m, _ in mid_refs]
    ref_map = dict(mid_refs)
    placeholders = ",".join("?" * len(mid_ids))
    cursor = conn.cursor()
    cursor.execute(f"""
        SELECT u_m.station_id, u_m.stop_time, u_b.stop_time, u_m.train_num
        FROM train_stops u_m
        JOIN train_stops u_b
          ON u_b.train_num = u_m.train_num AND u_b.station_id = ?
        WHERE u_m.station_id IN ({placeholders})
          AND u_m.stop_no < u_b.stop_no
          AND u_m.train_num != ?
        ORDER BY u_m.stop_time
    """, [to_station_id] + mid_ids + [excluded_train_num])
    solutions = []
    for mid_id, u_m_depart, u_b_arrive, u_train in cursor.fetchall():
        ref_time = ref_map.get(mid_id)
        if ref_time and u_m_depart:
            gap = _calc_time_diff_minutes(ref_time, u_m_depart)
            if gap >= min_gap:
                solutions.append({
                    "mid_id": mid_id,
                    "train_num": u_train,
                    "u_m_depart": u_m_depart,
                    "u_b_arrive": u_b_arrive,
                    "ref_time": ref_time,
                    "gap": gap,
                })
    return solutions


def check_pair(conn: sqlite3.Connection, from_id: str, to_id: str,
               min_gap: int = 20, max_attempts: int = 15) -> dict:
    """
    判定站对 (A→B)：
    - direct: 是否有直达车
    - transfer: 是否能通过一次换乘实现（重试次数与自动出题一致：≤ min(len(T), 15) 辆 T）
    - direct_trains: 直达车次列表
    - sample_transfer: 示例换乘 {t_train, mid, u_train}（有则给出）
    """
    # --- 直达 ---
    direct_routes = get_routes_between(conn, from_id, to_id)
    direct_trains = [r["train_num"] for r in direct_routes] if direct_routes else []

    # --- 换乘 ---
    routes = _get_routes_for_transfer(conn, from_id, to_id)
    # 与自动出题一致：随机打乱 T 候选，最多尝试 min(len, 15) 辆
    random.shuffle(routes)
    max_attempts = min(len(routes), max_attempts)
    sample_transfer = None
    transfer_ok = False

    # 目标站城市（用于过滤同城中间站）
    cursor = conn.cursor()
    cursor.execute("SELECT station_name FROM stations WHERE station_id = ?", (to_id,))
    row = cursor.fetchone()
    dest_name = row[0] if row else to_id
    dest_city = _get_city_key(dest_name)

    for attempt in range(max_attempts):
        t = routes[attempt]
        stops = t["stops"]
        from_idx = t["from_idx"]
        # 中间站：A 之后的所有站（T 不经过 B，故终点即末站）
        middle_indices = [i for i in range(from_idx + 1, len(stops))]
        # 与自动出题一致：过滤与 B 同城的中间站
        middle_indices = [
            i for i in middle_indices
            if _get_city_key(stops[i]["station_name"]) != dest_city
        ]
        if not middle_indices:
            continue
        mid_refs = [(stops[i]["station_id"], stops[i]["stop_time"]) for i in middle_indices]
        solutions = _find_transfer_solutions(conn, t["train_num"], mid_refs, to_id, min_gap)
        if solutions:
            transfer_ok = True
            s = solutions[0]
            sample_transfer = {
                "t_train": t["train_num"],
                "mid": s["mid_id"],
                "u_train": s["train_num"],
            }
            break

    return {
        "direct": len(direct_trains) > 0,
        "transfer": transfer_ok,
        "direct_trains": direct_trains,
        "sample_transfer": sample_transfer,
    }


# ---------------------------------------------------------------------------
# 读取站对表 / 写出结果
# ---------------------------------------------------------------------------

def read_pairs(path: str) -> list:
    """读取站对表（同 2.xlsx 格式：每行两列[出发站, 到达站]），跳过空行与同站对。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    pairs = []
    for row in wb.active.iter_rows(values_only=True):
        if not row:
            continue
        a = (str(row[0]).strip() if row[0] is not None else "")
        b = (str(row[1]).strip() if len(row) > 1 and row[1] is not None else "")
        if a and b and a != b:
            pairs.append([a, b])
    wb.close()
    return pairs


def station_display_name(conn: sqlite3.Connection, station_id: str) -> str:
    cursor = conn.cursor()
    cursor.execute("SELECT station_name FROM stations WHERE station_id = ?", (station_id,))
    row = cursor.fetchone()
    return row[0] if row else station_id


def write_result(conn: sqlite3.Connection, pairs: list, results: list, out_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "站对检查"
    ws.append(["出发站", "到达站", "是否有直达车", "是否能通过换乘实现",
               "直达车次列表", "示例换乘方案", "备注"])
    for (a, b), res in zip(pairs, results):
        if res is None:
            ws.append([a, b, "-", "-", "", "", "站对解析失败（站不存在）"])
            continue
        sample = ""
        if res["sample_transfer"]:
            s = res["sample_transfer"]
            mid_name = station_display_name(conn, s["mid"])
            sample = f"{s['t_train']} {a}→{mid_name} 换 {s['u_train']} {mid_name}→{b}"
        ws.append([
            a, b,
            "是" if res["direct"] else "否",
            "是" if res["transfer"] else "否",
            "、".join(res["direct_trains"]),
            sample,
            "",
        ])
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    for col_idx, width in [(1, 14), (2, 14), (3, 14), (4, 18), (5, 42), (6, 42), (7, 22)]:
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    wb.save(out_path)


# ---------------------------------------------------------------------------
# 主入口
# ---------------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(description="站对检查：是否有直达车 / 是否能通过换乘实现")
    parser.add_argument("input", help="输入站对表 Excel（格式同 2.xlsx：每行[出发站, 到达站]）")
    parser.add_argument("-o", "--output", default=None,
                        help="输出 Excel 路径（默认 = assist/station_pair_check_out.xlsx）")
    parser.add_argument("--db", default=None, help="基础数据库路径（默认 data/railway.db）")
    parser.add_argument("--min-gap", type=int, default=20, help="换乘衔接最短分钟（默认 20，与自动出题一致）")
    parser.add_argument("--max-attempts", type=int, default=15, help="候选首程车 T 的最大重试数（默认 15，与自动出题一致）")
    args = parser.parse_args()

    db_path = args.db or RAILWAY_DB_PATH
    out_path = args.output or os.path.join(os.path.dirname(os.path.abspath(__file__)), "station_pair_check_out.xlsx")

    if not os.path.exists(db_path):
        print(f"[错误] 基础数据库不存在: {db_path}，请先运行 collector.py 采集数据", file=sys.stderr)
        return 1
    if not os.path.exists(args.input):
        print(f"[错误] 输入文件不存在: {args.input}", file=sys.stderr)
        return 1

    conn = sqlite3.connect(db_path)
    try:
        pairs = read_pairs(args.input)
        if not pairs:
            print("[警告] 输入中没有有效的站对（每行两列[出发站, 到达站]，跳过空行与同站对）", file=sys.stderr)
        results = []
        for a, b in pairs:
            try:
                from_id = resolve_station_name_or_id(conn, a)
                to_id = resolve_station_name_or_id(conn, b)
                if not from_id or not to_id or from_id == to_id:
                    results.append(None)
                    continue
                # 站不在库中（resolve 返回原文）时视为解析失败
                cursor = conn.cursor()
                cursor.execute("SELECT 1 FROM stations WHERE station_id = ?", (from_id,))
                if cursor.fetchone() is None:
                    results.append(None)
                    continue
                cursor.execute("SELECT 1 FROM stations WHERE station_id = ?", (to_id,))
                if cursor.fetchone() is None:
                    results.append(None)
                    continue
                results.append(check_pair(conn, from_id, to_id, args.min_gap, args.max_attempts))
            except Exception as e:
                results.append(None)
        write_result(conn, pairs, results, out_path)
        print(f"[完成] 共检查 {len(pairs)} 个站对，已导出: {out_path}")
    finally:
        conn.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())